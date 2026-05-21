#!/usr/bin/env python3
"""
Hospital Triage System - Interactive CLI Interface
Bệnh Viện Đa Khoa - Hệ thống xếp hạng khám bệnh
"""

import json
import csv
import os
import re
import shutil
import sys
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Set
from pathlib import Path

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich.prompt import Prompt, Confirm, IntPrompt
from rich.progress import Progress, SpinnerColumn, TextColumn
from rich.layout import Layout
from rich.columns import Columns
from rich import box

from hospital_triage import (
    VirtualClock,
    Patient,
    Doctor,
    Department,
    HospitalTriageSystem,
    BillingRecord,
)
from billing_system import BillingSystem, Bill

console = Console()


class DataPersistence:
    """Handle data persistence with auto-save, load, and backup."""

    def __init__(self, data_dir: str = "data"):
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(exist_ok=True)
        self.backup_dir = self.data_dir / "backups"
        self.backup_dir.mkdir(exist_ok=True)
        self.patients_file = self.data_dir / "patients.json"
        self.doctors_file = self.data_dir / "doctors.json"
        self.departments_file = self.data_dir / "departments.json"
        self.appointments_file = self.data_dir / "appointments.json"
        self.bills_file = self.data_dir / "bills.json"
        self.system_state_file = self.data_dir / "system_state.json"

    def _create_backup(self, filepath: Path):
        """Create a timestamped backup of a file."""
        if filepath.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{filepath.stem}_{timestamp}{filepath.suffix}"
            backup_path = self.backup_dir / backup_name
            shutil.copy2(filepath, backup_path)
            # Keep only last 10 backups
            backups = sorted(self.backup_dir.glob(f"{filepath.stem}_*"))
            for old_backup in backups[:-10]:
                old_backup.unlink()

    def save_data(self, cli_system):
        """Save all system data to JSON files."""
        try:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console,
            ) as progress:
                task = progress.add_task("Saving data...", total=None)

                # Save patients
                self._create_backup(self.patients_file)
                patients_data = {}
                for pid, patient in cli_system.system.patients.items():
                    patients_data[pid] = {
                        "id": patient.id,
                        "name": patient.name,
                        "priority": patient.priority,
                        "checked_in_at": (
                            patient.checked_in_at.isoformat()
                            if patient.checked_in_at
                            else None
                        ),
                        "department": patient.department,
                        "doctor_id": patient.doctor_id,
                        "status": patient.status,
                        "appointment_time": (
                            patient.appointment_time.isoformat()
                            if patient.appointment_time
                            else None
                        ),
                        "is_walk_in": patient.is_walk_in,
                    }
                self.patients_file.write_text(
                    json.dumps(patients_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

                # Save doctors
                self._create_backup(self.doctors_file)
                doctors_data = {}
                for dept in cli_system.system.departments.values():
                    for doctor in dept.doctors:
                        doctors_data[doctor.id] = {
                            "id": doctor.id,
                            "name": doctor.name,
                            "department": doctor.department,
                            "max_slots_per_hour": doctor.max_slots_per_hour,
                            "current_patient_id": (
                                doctor.current_patient.id
                                if doctor.current_patient
                                else None
                            ),
                        }
                self.doctors_file.write_text(
                    json.dumps(doctors_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

                # Save departments
                self._create_backup(self.departments_file)
                depts_data = {}
                for name, dept in cli_system.system.departments.items():
                    depts_data[name] = {
                        "name": dept.name,
                        "doctor_ids": [d.id for d in dept.doctors],
                        "queue_ids": [p.id for p in dept.queue],
                    }
                self.departments_file.write_text(
                    json.dumps(depts_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

                # Save system state (clock, counters)
                self._create_backup(self.system_state_file)
                state_data = {
                    "clock_time": cli_system.system.clock.get_time().isoformat(),
                    "patient_counter": cli_system.patient_counter,
                    "doctor_counter": cli_system.doctor_counter,
                }
                self.system_state_file.write_text(
                    json.dumps(state_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

                # Save billing system
                self._create_backup(self.bills_file)
                bills_data = {
                    "bills": {},
                    "patient_bills": cli_system.billing_system.patient_bills,
                    "service_catalog": cli_system.billing_system.service_catalog,
                    "bill_counter": cli_system.billing_system._bill_counter,
                }
                for bill_id, bill in cli_system.billing_system.bills.items():
                    bills_data["bills"][bill_id] = {
                        "bill_id": bill.bill_id,
                        "patient_id": bill.patient_id,
                        "doctor_id": bill.doctor_id,
                        "department_id": bill.department_id,
                        "items": bill.items,
                        "total_amount": bill.total_amount,
                        "is_paid": bill.is_paid,
                        "created_at": bill.created_at.isoformat(),
                    }
                self.bills_file.write_text(
                    json.dumps(bills_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

                progress.update(task, completed=True)

            console.print("[bold green]Dữ liệu đã được lưu thành công![/bold green]")
            return True
        except Exception as e:
            console.print(f"[bold red]Lỗi khi lưu dữ liệu: {e}[/bold red]")
            return False

    def load_data(self, cli_system):
        """Load all system data from JSON files."""
        try:
            if not any(
                [
                    self.patients_file.exists(),
                    self.doctors_file.exists(),
                    self.departments_file.exists(),
                ]
            ):
                console.print(
                    "[bold yellow]Không tìm thấy dữ liệu cũ. Khởi tạo hệ thống mới.[/bold yellow]"
                )
                return False

            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console,
            ) as progress:
                task = progress.add_task("Loading data...", total=None)

                # Load system state
                if self.system_state_file.exists():
                    state_data = json.loads(
                        self.system_state_file.read_text(encoding="utf-8")
                    )
                    clock_time = datetime.fromisoformat(state_data["clock_time"])
                    cli_system.system.clock = VirtualClock(clock_time)
                    cli_system.patient_counter = state_data.get("patient_counter", 0)
                    cli_system.doctor_counter = state_data.get("doctor_counter", 0)

                # Load patients
                if self.patients_file.exists():
                    patients_data = json.loads(
                        self.patients_file.read_text(encoding="utf-8")
                    )
                    for pid, pdata in patients_data.items():
                        patient = Patient(
                            id=pdata["id"],
                            name=pdata["name"],
                            priority=pdata["priority"],
                        )
                        if pdata.get("checked_in_at"):
                            patient.checked_in_at = datetime.fromisoformat(
                                pdata["checked_in_at"]
                            )
                        patient.department = pdata.get("department")
                        patient.doctor_id = pdata.get("doctor_id")
                        patient.status = pdata.get("status", "waiting")
                        if pdata.get("appointment_time"):
                            patient.appointment_time = datetime.fromisoformat(
                                pdata["appointment_time"]
                            )
                        patient.is_walk_in = pdata.get("is_walk_in", False)
                        cli_system.system.patients[pid] = patient

                # Load departments
                if self.departments_file.exists():
                    depts_data = json.loads(
                        self.departments_file.read_text(encoding="utf-8")
                    )
                    for name, ddata in depts_data.items():
                        dept = Department(name)
                        cli_system.system.departments[name] = dept

                # Load doctors
                if self.doctors_file.exists():
                    doctors_data = json.loads(
                        self.doctors_file.read_text(encoding="utf-8")
                    )
                    for did, ddata in doctors_data.items():
                        doctor = Doctor(
                            id=ddata["id"],
                            name=ddata["name"],
                            department=ddata["department"],
                            max_slots_per_hour=ddata.get("max_slots_per_hour", 4),
                        )
                        if ddata.get("current_patient_id"):
                            doctor.current_patient = cli_system.system.patients.get(
                                ddata["current_patient_id"]
                            )
                        if doctor.department in cli_system.system.departments:
                            cli_system.system.departments[doctor.department].add_doctor(
                                doctor
                            )

                # Restore department queues
                if self.departments_file.exists():
                    depts_data = json.loads(
                        self.departments_file.read_text(encoding="utf-8")
                    )
                    for name, ddata in depts_data.items():
                        dept = cli_system.system.departments.get(name)
                        if dept:
                            for pid in ddata.get("queue_ids", []):
                                patient = cli_system.system.patients.get(pid)
                                if patient:
                                    dept.queue.append(patient)

                # Load billing system
                if self.bills_file.exists():
                    bills_data = json.loads(self.bills_file.read_text(encoding="utf-8"))
                    cli_system.billing_system.patient_bills = bills_data.get(
                        "patient_bills", {}
                    )
                    cli_system.billing_system.service_catalog = bills_data.get(
                        "service_catalog", BillingSystem.SERVICE_CATALOG
                    )
                    cli_system.billing_system._bill_counter = bills_data.get(
                        "bill_counter", 0
                    )
                    for bill_id, bdata in bills_data.get("bills", {}).items():
                        bill = Bill(
                            bill_id=bdata["bill_id"],
                            patient_id=bdata["patient_id"],
                            doctor_id=bdata["doctor_id"],
                            department_id=bdata["department_id"],
                        )
                        bill.items = bdata.get("items", [])
                        bill.total_amount = bdata.get("total_amount", 0.0)
                        bill.is_paid = bdata.get("is_paid", False)
                        if bdata.get("created_at"):
                            bill.created_at = datetime.fromisoformat(
                                bdata["created_at"]
                            )
                        cli_system.billing_system.bills[bill_id] = bill

                progress.update(task, completed=True)

            console.print("[bold green]Dữ liệu đã được tải thành công![/bold green]")
            return True
        except Exception as e:
            console.print(f"[bold red]Lỗi khi tải dữ liệu: {e}[/bold red]")
            return False


class Validation:
    """Input validation utilities."""

    VIETNAMESE_PHONE_PATTERN = re.compile(
        r"^(0[3|5|7|8|9][0-9]{8}|\+84[3|5|7|8|9][0-9]{8})$"
    )
    EMAIL_PATTERN = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

    @staticmethod
    def validate_phone(phone: str) -> bool:
        """Validate Vietnamese phone number format."""
        cleaned = phone.replace(" ", "").replace("-", "").replace(".", "")
        return bool(Validation.VIETNAMESE_PHONE_PATTERN.match(cleaned))

    @staticmethod
    def validate_date(date_str: str) -> Optional[datetime]:
        """Validate date string (DD/MM/YYYY or YYYY-MM-DD)."""
        formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def validate_time_slot(time_str: str) -> bool:
        """Validate time slot format (HH:MM)."""
        try:
            datetime.strptime(time_str, "%H:%M")
            return True
        except ValueError:
            return False

    @staticmethod
    def validate_not_empty(value: str, field_name: str) -> str:
        """Validate that a field is not empty."""
        if not value or not value.strip():
            raise ValueError(f"{field_name} không được để trống")
        return value.strip()

    @staticmethod
    def validate_age(age_str: str) -> int:
        """Validate age input."""
        try:
            age = int(age_str)
            if age < 0 or age > 150:
                raise ValueError("Tuổi phải từ 0 đến 150")
            return age
        except ValueError as e:
            if "invalid literal" in str(e):
                raise ValueError("Tuổi phải là số nguyên")
            raise

    @staticmethod
    def validate_priority(priority_str: str) -> int:
        """Validate priority level (1-4)."""
        try:
            priority = int(priority_str)
            if priority < 1 or priority > 4:
                raise ValueError("Mức độ ưu tiên phải từ 1 đến 4")
            return priority
        except ValueError as e:
            if "invalid literal" in str(e):
                raise ValueError("Mức độ ưu tiên phải là số nguyên")
            raise


class HospitalCLI:
    """Interactive CLI for Hospital Triage System."""

    PRIORITY_NAMES = {
        1: "[bold red]CẤP CỨU[/bold red]",
        2: "[bold yellow]ƯU TIÊN[/bold yellow]",
        3: "[bold blue]BÌNH THƯỜNG[/bold blue]",
        4: "[bold green]THẤP[/bold green]",
    }

    STATUS_COLORS = {
        "waiting": "[yellow]Đang chờ[/yellow]",
        "in_exam": "[blue]Đang khám[/blue]",
        "completed": "[green]Hoàn thành[/green]",
        "billed": "[cyan]Đã tính tiền[/cyan]",
    }

    def __init__(self):
        self.clock = VirtualClock(datetime(2024, 1, 15, 8, 0, 0))
        self.system = HospitalTriageSystem(self.clock)
        self.billing_system = BillingSystem()
        self.persistence = DataPersistence()
        self.patient_counter = 0
        self.doctor_counter = 0
        self.running = True

        # Initialize default departments
        self._init_default_departments()

        # Try to load existing data
        self.persistence.load_data(self)

    def _init_default_departments(self):
        """Initialize default departments."""
        default_departments = [
            "Cấp cứu",
            "Nội tổng quát",
            "Ngoại khoa",
            "Nhi khoa",
            "Sản khoa",
            "Tai mũi họng",
            "Mắt",
            "Da liễu",
            "Tim mạch",
            "Thần kinh",
        ]
        for dept_name in default_departments:
            if dept_name not in self.system.departments:
                self.system.add_department(dept_name)

    def display_main_menu(self):
        """Display the main menu with Rich styling."""
        console.print()
        title = Text("HOSPITAL TRIAGE SYSTEM v2.0", style="bold cyan")
        subtitle = Text("Bệnh Viện Đa Khoa", style="bold white")
        header = Panel(
            Text.assemble(title, "\n", subtitle),
            box=box.DOUBLE,
            border_style="bright_blue",
            padding=(1, 2),
        )
        console.print(header)
        console.print()

        menu_items = [
            ("[1]", "Quản lý Bệnh nhân", "bright_green"),
            ("[2]", "Quản lý Bác sĩ", "bright_green"),
            ("[3]", "Quản lý Khoa", "bright_green"),
            ("[4]", "Đặt lịch khám", "bright_yellow"),
            ("[5]", "Check-in bệnh nhân", "bright_yellow"),
            ("[6]", "Bác sĩ khám bệnh", "bright_yellow"),
            ("[7]", "Xem Dashboard", "bright_cyan"),
            ("[8]", "Tính tiền / Hóa đơn", "bright_magenta"),
            ("[9]", "Báo cáo & Thống kê", "bright_magenta"),
            ("[10]", "Sinh dữ liệu mẫu", "bright_white"),
            ("[0]", "Thoát", "bold red"),
        ]

        for num, label, color in menu_items:
            console.print(f"  {num} ", style="bold", end="")
            console.print(label, style=color)

        console.print()

    def get_menu_choice(
        self, prompt_text: str = "Chọn chức năng", max_choice: int = 10
    ) -> int:
        """Get validated menu choice from user."""
        while True:
            try:
                choice = console.input(f"[bold]{prompt_text} (0-{max_choice}): [/bold]")
                choice = int(choice)
                if 0 <= choice <= max_choice:
                    return choice
                console.print(
                    f"[bold red]Vui lòng chọn từ 0 đến {max_choice}[/bold red]"
                )
            except ValueError:
                console.print("[bold red]Vui lòng nhập số[/bold red]")

    def pause(self):
        """Pause for user to continue."""
        console.input("\n[dim]Nhấn Enter để tiếp tục...[/dim]")

    def clear_screen(self):
        """Clear the console screen."""
        console.clear()

    # ═══════════════════════════════════════════════════════════════
    # PATIENT MANAGEMENT
    # ═══════════════════════════════════════════════════════════════

    def patient_menu(self):
        """Patient management sub-menu."""
        while True:
            self.clear_screen()
            console.print(Panel("[bold]QUẢN LÝ BỆNH NHÂN[/bold]", border_style="green"))
            console.print("  [1] Thêm bệnh nhân mới")
            console.print("  [2] Xem danh sách bệnh nhân")
            console.print("  [3] Xem chi tiết bệnh nhân")
            console.print("  [4] Sửa thông tin bệnh nhân")
            console.print("  [5] Xóa bệnh nhân")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 5)

            if choice == 1:
                self.add_patient()
            elif choice == 2:
                self.view_patients()
            elif choice == 3:
                self.view_patient_detail()
            elif choice == 4:
                self.edit_patient()
            elif choice == 5:
                self.delete_patient()
            elif choice == 0:
                break

            self.pause()

    def add_patient(self):
        """Add a new patient with validation."""
        console.print(Panel("[bold green]THÊM BỆNH NHÂN MỚI[/bold green]"))

        try:
            name = console.input("[bold]Họ và tên:[/bold] ")
            Validation.validate_not_empty(name, "Họ và tên")

            age_str = console.input("[bold]Tuổi:[/bold] ")
            age = Validation.validate_age(age_str)

            symptoms = console.input("[bold]Triệu chứng:[/bold] ")

            console.print("\n[bold]Mức độ ưu tiên:[/bold]")
            console.print("  1 - Cấp cứu (Emergency)")
            console.print("  2 - Ưu tiên (Urgent)")
            console.print("  3 - Bình thường (Normal)")
            console.print("  4 - Thấp (Low)")
            priority_str = console.input("[bold]Chọn (1-4, mặc định 3):[/bold] ") or "3"
            priority = Validation.validate_priority(priority_str)

            phone = console.input("[bold]Số điện thoại:[/bold] ")
            if phone and not Validation.validate_phone(phone):
                console.print(
                    "[bold yellow]Cảnh báo: Số điện thoại không đúng định dạng VN[/bold yellow]"
                )

            address = console.input("[bold]Địa chỉ:[/bold] ")

            # Generate patient ID
            self.patient_counter += 1
            patient_id = f"BN{self.patient_counter:04d}"

            patient = Patient(id=patient_id, name=name, priority=priority)
            self.system.register_patient(patient)

            # Store additional info
            patient.age = age
            patient.symptoms = symptoms
            patient.phone = phone
            patient.address = address

            console.print(
                f"\n[bold green]Đã thêm bệnh nhân {patient_id} - {name}[/bold green]"
            )

        except ValueError as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Lỗi không mong muốn: {e}[/bold red]")

    def view_patients(self):
        """Display all patients in a table."""
        if not self.system.patients:
            console.print("[bold yellow]Chưa có bệnh nhân nào[/bold yellow]")
            return

        table = Table(
            title="DANH SÁCH BỆNH NHÂN",
            box=box.ROUNDED,
            header_style="bold cyan",
            border_style="cyan",
        )
        table.add_column("Mã BN", style="bold", width=10)
        table.add_column("Họ tên", width=25)
        table.add_column("Tuổi", justify="center", width=6)
        table.add_column("Ưu tiên", justify="center", width=12)
        table.add_column("Trạng thái", justify="center", width=14)
        table.add_column("Khoa", width=18)
        table.add_column("Bác sĩ", width=12)

        for patient in self.system.patients.values():
            priority_display = self.PRIORITY_NAMES.get(
                patient.priority, str(patient.priority)
            )
            status_display = self.STATUS_COLORS.get(patient.status, patient.status)
            age = getattr(patient, "age", "N/A")
            dept = patient.department or "Chưa phân"
            doctor = patient.doctor_id or "Chưa có"

            table.add_row(
                patient.id,
                patient.name,
                str(age),
                priority_display,
                status_display,
                dept,
                doctor,
            )

        console.print(table)
        console.print(f"\n[dim]Tổng số: {len(self.system.patients)} bệnh nhân[/dim]")

    def view_patient_detail(self):
        """View detailed information for a specific patient."""
        patient_id = console.input("[bold]Nhập mã bệnh nhân:[/bold] ")
        patient = self.system.patients.get(patient_id)

        if not patient:
            console.print(f"[bold red]Không tìm thấy bệnh nhân {patient_id}[/bold red]")
            return

        info = []
        info.append(f"[bold]Mã bệnh nhân:[/bold] {patient.id}")
        info.append(f"[bold]Họ tên:[/bold] {patient.name}")
        info.append(f"[bold]Tuổi:[/bold] {getattr(patient, 'age', 'N/A')}")
        info.append(
            f"[bold]Mức độ ưu tiên:[/bold] {self.PRIORITY_NAMES.get(patient.priority, str(patient.priority))}"
        )
        info.append(
            f"[bold]Trạng thái:[/bold] {self.STATUS_COLORS.get(patient.status, patient.status)}"
        )
        info.append(f"[bold]Triệu chứng:[/bold] {getattr(patient, 'symptoms', 'N/A')}")
        info.append(f"[bold]Số điện thoại:[/bold] {getattr(patient, 'phone', 'N/A')}")
        info.append(f"[bold]Địa chỉ:[/bold] {getattr(patient, 'address', 'N/A')}")
        info.append(f"[bold]Khoa:[/bold] {patient.department or 'Chưa phân'}")
        info.append(f"[bold]Bác sĩ:[/bold] {patient.doctor_id or 'Chưa có'}")
        if patient.appointment_time:
            info.append(
                f"[bold]Lịch hẹn:[/bold] {patient.appointment_time.strftime('%d/%m/%Y %H:%M')}"
            )
        if patient.checked_in_at:
            info.append(
                f"[bold]Check-in:[/bold] {patient.checked_in_at.strftime('%d/%m/%Y %H:%M')}"
            )

        panel = Panel(
            "\n".join(info),
            title=f"[bold]THÔNG TIN BỆNH NHÂN {patient.id}[/bold]",
            border_style="green",
            box=box.ROUNDED,
        )
        console.print(panel)

    def edit_patient(self):
        """Edit patient information."""
        patient_id = console.input("[bold]Nhập mã bệnh nhân cần sửa:[/bold] ")
        patient = self.system.patients.get(patient_id)

        if not patient:
            console.print(f"[bold red]Không tìm thấy bệnh nhân {patient_id}[/bold red]")
            return

        console.print(f"\n[bold]Sửa thông tin cho {patient.name} ({patient.id})[/bold]")
        console.print("[dim]Để trống để giữ nguyên giá trị cũ[/dim]\n")

        new_name = console.input(f"[bold]Họ tên ({patient.name}):[/bold] ")
        if new_name.strip():
            patient.name = Validation.validate_not_empty(new_name, "Họ và tên")

        new_age = console.input(
            f"[bold]Tuổi ({getattr(patient, 'age', 'N/A')}):[/bold] "
        )
        if new_age.strip():
            patient.age = Validation.validate_age(new_age)

        new_symptoms = console.input(
            f"[bold]Triệu chứng ({getattr(patient, 'symptoms', 'N/A')}):[/bold] "
        )
        if new_symptoms.strip():
            patient.symptoms = new_symptoms

        new_phone = console.input(
            f"[bold]SĐT ({getattr(patient, 'phone', 'N/A')}):[/bold] "
        )
        if new_phone.strip():
            if not Validation.validate_phone(new_phone):
                console.print(
                    "[bold yellow]Cảnh báo: Số điện thoại không đúng định dạng[/bold yellow]"
                )
            patient.phone = new_phone

        new_address = console.input(
            f"[bold]Địa chỉ ({getattr(patient, 'address', 'N/A')}):[/bold] "
        )
        if new_address.strip():
            patient.address = new_address

        console.print(
            f"\n[bold green]Đã cập nhật thông tin bệnh nhân {patient_id}[/bold green]"
        )

    def delete_patient(self):
        """Delete a patient from the system."""
        patient_id = console.input("[bold]Nhập mã bệnh nhân cần xóa:[/bold] ")
        patient = self.system.patients.get(patient_id)

        if not patient:
            console.print(f"[bold red]Không tìm thấy bệnh nhân {patient_id}[/bold red]")
            return

        console.print(
            f"\n[bold yellow]Bạn có chắc muốn xóa bệnh nhân {patient.name} ({patient_id})?[/bold yellow]"
        )
        confirm = Confirm.ask("Xác nhận xóa", default=False)

        if confirm:
            del self.system.patients[patient_id]
            console.print(f"[bold green]Đã xóa bệnh nhân {patient_id}[/bold green]")
        else:
            console.print("[dim]Đã hủy thao tác xóa[/dim]")

    # ═══════════════════════════════════════════════════════════════
    # DOCTOR MANAGEMENT
    # ═══════════════════════════════════════════════════════════════

    def doctor_menu(self):
        """Doctor management sub-menu."""
        while True:
            self.clear_screen()
            console.print(Panel("[bold]QUẢN LÝ BÁC SĨ[/bold]", border_style="green"))
            console.print("  [1] Thêm bác sĩ mới")
            console.print("  [2] Xem danh sách bác sĩ")
            console.print("  [3] Xem chi tiết bác sĩ")
            console.print("  [4] Sửa thông tin bác sĩ")
            console.print("  [5] Xóa bác sĩ")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 5)

            if choice == 1:
                self.add_doctor()
            elif choice == 2:
                self.view_doctors()
            elif choice == 3:
                self.view_doctor_detail()
            elif choice == 4:
                self.edit_doctor()
            elif choice == 5:
                self.delete_doctor()
            elif choice == 0:
                break

            self.pause()

    def add_doctor(self):
        """Add a new doctor."""
        console.print(Panel("[bold green]THÊM BÁC SĨ MỚI[/bold green]"))

        try:
            name = console.input("[bold]Họ và tên:[/bold] ")
            Validation.validate_not_empty(name, "Họ và tên")

            console.print("\n[bold]Danh sách khoa:[/bold]")
            for i, dept_name in enumerate(self.system.departments.keys(), 1):
                console.print(f"  {i}. {dept_name}")

            dept_choice = console.input("[bold]Chọn khoa (số):[/bold] ")
            dept_list = list(self.system.departments.keys())
            dept_idx = int(dept_choice) - 1
            if dept_idx < 0 or dept_idx >= len(dept_list):
                raise ValueError("Khoa không hợp lệ")
            department = dept_list[dept_idx]

            specialty = console.input("[bold]Chuyên khoa:[/bold] ")
            phone = console.input("[bold]Số điện thoại:[/bold] ")
            if phone and not Validation.validate_phone(phone):
                console.print(
                    "[bold yellow]Cảnh báo: Số điện thoại không đúng định dạng[/bold yellow]"
                )

            self.doctor_counter += 1
            doctor_id = f"BS{self.doctor_counter:04d}"

            doctor = Doctor(
                id=doctor_id,
                name=name,
                department=department,
            )
            self.system.add_doctor(doctor)
            doctor.specialty = specialty
            doctor.phone = phone

            console.print(
                f"\n[bold green]Đã thêm bác sĩ {doctor_id} - {name}[/bold green]"
            )

        except ValueError as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")

    def view_doctors(self):
        """Display all doctors in a table."""
        all_doctors = []
        for dept in self.system.departments.values():
            all_doctors.extend(dept.doctors)

        if not all_doctors:
            console.print("[bold yellow]Chưa có bác sĩ nào[/bold yellow]")
            return

        table = Table(
            title="DANH SÁCH BÁC SĨ",
            box=box.ROUNDED,
            header_style="bold cyan",
            border_style="cyan",
        )
        table.add_column("Mã BS", style="bold", width=10)
        table.add_column("Họ tên", width=25)
        table.add_column("Khoa", width=18)
        table.add_column("Trạng thái", justify="center", width=14)
        table.add_column("BN hiện tại", width=14)
        table.add_column("Lịch hẹn", justify="center", width=10)

        for doctor in all_doctors:
            if doctor.current_patient:
                status = "[red]Đang khám[/red]"
                current_patient = doctor.current_patient.name
            else:
                status = "[green]Rảnh[/green]"
                current_patient = "-"

            appointment_count = doctor.get_appointment_count()

            table.add_row(
                doctor.id,
                doctor.name,
                doctor.department,
                status,
                current_patient,
                str(appointment_count),
            )

        console.print(table)
        console.print(f"\n[dim]Tổng số: {len(all_doctors)} bác sĩ[/dim]")

    def view_doctor_detail(self):
        """View detailed information for a specific doctor."""
        doctor_id = console.input("[bold]Nhập mã bác sĩ:[/bold] ")

        doctor = None
        for dept in self.system.departments.values():
            for d in dept.doctors:
                if d.id == doctor_id:
                    doctor = d
                    break
            if doctor:
                break

        if not doctor:
            console.print(f"[bold red]Không tìm thấy bác sĩ {doctor_id}[/bold red]")
            return

        info = []
        info.append(f"[bold]Mã bác sĩ:[/bold] {doctor.id}")
        info.append(f"[bold]Họ tên:[/bold] {doctor.name}")
        info.append(f"[bold]Khoa:[/bold] {doctor.department}")
        info.append(f"[bold]Chuyên khoa:[/bold] {getattr(doctor, 'specialty', 'N/A')}")
        info.append(f"[bold]Số điện thoại:[/bold] {getattr(doctor, 'phone', 'N/A')}")
        if doctor.current_patient:
            info.append(
                f"[bold]BN đang khám:[/bold] {doctor.current_patient.name} ({doctor.current_patient.id})"
            )
        else:
            info.append("[bold]Trạng thái:[/bold] [green]Đang rảnh[/green]")

        info.append(f"[bold]Số lịch hẹn:[/bold] {doctor.get_appointment_count()}")
        info.append(f"[bold]Số BN đã khám:[/bold] {len(doctor.completed_patients)}")

        panel = Panel(
            "\n".join(info),
            title=f"[bold]THÔNG TIN BÁC SĨ {doctor.id}[/bold]",
            border_style="green",
            box=box.ROUNDED,
        )
        console.print(panel)

    def edit_doctor(self):
        """Edit doctor information."""
        doctor_id = console.input("[bold]Nhập mã bác sĩ cần sửa:[/bold] ")

        doctor = None
        for dept in self.system.departments.values():
            for d in dept.doctors:
                if d.id == doctor_id:
                    doctor = d
                    break
            if doctor:
                break

        if not doctor:
            console.print(f"[bold red]Không tìm thấy bác sĩ {doctor_id}[/bold red]")
            return

        console.print(f"\n[bold]Sửa thông tin cho {doctor.name} ({doctor.id})[/bold]")
        console.print("[dim]Để trống để giữ nguyên giá trị cũ[/dim]\n")

        new_name = console.input(f"[bold]Họ tên ({doctor.name}):[/bold] ")
        if new_name.strip():
            doctor.name = Validation.validate_not_empty(new_name, "Họ và tên")

        new_specialty = console.input(
            f"[bold]Chuyên khoa ({getattr(doctor, 'specialty', 'N/A')}):[/bold] "
        )
        if new_specialty.strip():
            doctor.specialty = new_specialty

        new_phone = console.input(
            f"[bold]SĐT ({getattr(doctor, 'phone', 'N/A')}):[/bold] "
        )
        if new_phone.strip():
            if not Validation.validate_phone(new_phone):
                console.print(
                    "[bold yellow]Cảnh báo: Số điện thoại không đúng định dạng[/bold yellow]"
                )
            doctor.phone = new_phone

        console.print(
            f"\n[bold green]Đã cập nhật thông tin bác sĩ {doctor_id}[/bold green]"
        )

    def delete_doctor(self):
        """Delete a doctor from the system."""
        doctor_id = console.input("[bold]Nhập mã bác sĩ cần xóa:[/bold] ")

        doctor = None
        doctor_dept = None
        for dept in self.system.departments.values():
            for d in dept.doctors:
                if d.id == doctor_id:
                    doctor = d
                    doctor_dept = dept
                    break
            if doctor:
                break

        if not doctor:
            console.print(f"[bold red]Không tìm thấy bác sĩ {doctor_id}[/bold red]")
            return

        if doctor.current_patient:
            console.print("[bold red]Không thể xóa bác sĩ đang khám bệnh![/bold red]")
            return

        console.print(
            f"\n[bold yellow]Bạn có chắc muốn xóa bác sĩ {doctor.name} ({doctor_id})?[/bold yellow]"
        )
        confirm = Confirm.ask("Xác nhận xóa", default=False)

        if confirm and doctor_dept:
            doctor_dept.doctors.remove(doctor)
            console.print(f"[bold green]Đã xóa bác sĩ {doctor_id}[/bold green]")
        else:
            console.print("[dim]Đã hủy thao tác xóa[/dim]")

    # ═══════════════════════════════════════════════════════════════
    # DEPARTMENT MANAGEMENT
    # ═══════════════════════════════════════════════════════════════

    def department_menu(self):
        """Department management sub-menu."""
        while True:
            self.clear_screen()
            console.print(Panel("[bold]QUẢN LÝ KHOA[/bold]", border_style="green"))
            console.print("  [1] Xem danh sách khoa")
            console.print("  [2] Xem chi tiết khoa")
            console.print("  [3] Thêm khoa mới")
            console.print("  [4] Xóa khoa")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 4)

            if choice == 1:
                self.view_departments()
            elif choice == 2:
                self.view_department_detail()
            elif choice == 3:
                self.add_department()
            elif choice == 4:
                self.delete_department()
            elif choice == 0:
                break

            self.pause()

    def view_departments(self):
        """Display all departments in a table."""
        if not self.system.departments:
            console.print("[bold yellow]Chưa có khoa nào[/bold yellow]")
            return

        table = Table(
            title="DANH SÁCH KHOA",
            box=box.ROUNDED,
            header_style="bold cyan",
            border_style="cyan",
        )
        table.add_column("Tên khoa", style="bold", width=25)
        table.add_column("Số bác sĩ", justify="center", width=12)
        table.add_column("Số BN chờ", justify="center", width=12)
        table.add_column("BS đang rảnh", justify="center", width=14)

        for dept in self.system.departments.values():
            available = len(dept.get_available_doctors())
            table.add_row(
                dept.name,
                str(len(dept.doctors)),
                str(dept.get_queue_length()),
                f"[green]{available}[/green]",
            )

        console.print(table)

    def view_department_detail(self):
        """View detailed information for a specific department."""
        console.print("[bold]Danh sách khoa:[/bold]")
        for i, dept_name in enumerate(self.system.departments.keys(), 1):
            console.print(f"  {i}. {dept_name}")

        dept_choice = console.input("[bold]Chọn khoa (số):[/bold] ")
        dept_list = list(self.system.departments.keys())
        try:
            dept_idx = int(dept_choice) - 1
            if dept_idx < 0 or dept_idx >= len(dept_list):
                raise ValueError()
            dept_name = dept_list[dept_idx]
        except (ValueError, IndexError):
            console.print("[bold red]Khoa không hợp lệ[/bold red]")
            return

        dept = self.system.departments.get(dept_name)
        if not dept:
            console.print(f"[bold red]Không tìm thấy khoa {dept_name}[/bold red]")
            return

        info = []
        info.append(f"[bold]Tên khoa:[/bold] {dept.name}")
        info.append(f"[bold]Số bác sĩ:[/bold] {len(dept.doctors)}")
        info.append(f"[bold]Số bệnh nhân đang chờ:[/bold] {dept.get_queue_length()}")
        info.append(
            f"[bold]Bác sĩ đang rảnh:[/bold] {len(dept.get_available_doctors())}"
        )

        if dept.doctors:
            info.append("\n[bold]Danh sách bác sĩ:[/bold]")
            for doctor in dept.doctors:
                status = (
                    "[green]Rảnh[/green]"
                    if not doctor.current_patient
                    else "[red]Bận[/red]"
                )
                info.append(f"  - {doctor.name} ({doctor.id}) {status}")

        if dept.queue:
            info.append("\n[bold]Hàng đợi bệnh nhân:[/bold]")
            for i, patient in enumerate(dept.queue, 1):
                priority = self.PRIORITY_NAMES.get(
                    patient.priority, str(patient.priority)
                )
                info.append(f"  {i}. {patient.name} ({patient.id}) - {priority}")

        panel = Panel(
            "\n".join(info),
            title=f"[bold]CHI TIẾT KHOA {dept_name}[/bold]",
            border_style="green",
            box=box.ROUNDED,
        )
        console.print(panel)

    def add_department(self):
        """Add a new department."""
        dept_name = console.input("[bold]Tên khoa mới:[/bold] ")
        try:
            Validation.validate_not_empty(dept_name, "Tên khoa")
            if dept_name in self.system.departments:
                console.print(f"[bold red]Khoa {dept_name} đã tồn tại[/bold red]")
                return
            self.system.add_department(dept_name)
            console.print(f"[bold green]Đã thêm khoa {dept_name}[/bold green]")
        except ValueError as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")

    def delete_department(self):
        """Delete a department."""
        console.print("[bold]Danh sách khoa:[/bold]")
        for i, dept_name in enumerate(self.system.departments.keys(), 1):
            console.print(f"  {i}. {dept_name}")

        dept_choice = console.input("[bold]Chọn khoa cần xóa (số):[/bold] ")
        dept_list = list(self.system.departments.keys())
        try:
            dept_idx = int(dept_choice) - 1
            if dept_idx < 0 or dept_idx >= len(dept_list):
                raise ValueError()
            dept_name = dept_list[dept_idx]
        except (ValueError, IndexError):
            console.print("[bold red]Khoa không hợp lệ[/bold red]")
            return

        dept = self.system.departments.get(dept_name)
        if dept and (dept.doctors or dept.queue):
            console.print(
                "[bold red]Không thể xóa khoa đang có bác sĩ hoặc bệnh nhân![/bold red]"
            )
            return

        confirm = Confirm.ask(f"Xác nhận xóa khoa {dept_name}?", default=False)
        if confirm:
            del self.system.departments[dept_name]
            console.print(f"[bold green]Đã xóa khoa {dept_name}[/bold green]")

    # ═══════════════════════════════════════════════════════════════
    # APPOINTMENT BOOKING
    # ═══════════════════════════════════════════════════════════════

    def appointment_menu(self):
        """Appointment booking sub-menu."""
        while True:
            self.clear_screen()
            console.print(Panel("[bold]ĐẶT LỊCH KHÁM[/bold]", border_style="yellow"))
            console.print("  [1] Đặt lịch hẹn mới")
            console.print("  [2] Xem lịch hẹn theo bác sĩ")
            console.print("  [3] Xem lịch hẹn theo ngày")
            console.print("  [4] Hủy lịch hẹn")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 4)

            if choice == 1:
                self.book_appointment()
            elif choice == 2:
                self.view_appointments_by_doctor()
            elif choice == 3:
                self.view_appointments_by_date()
            elif choice == 4:
                self.cancel_appointment()
            elif choice == 0:
                break

            self.pause()

    def book_appointment(self):
        """Book an appointment with validation."""
        console.print(Panel("[bold yellow]ĐẶT LỊCH HẸN MỚI[/bold yellow]"))

        try:
            # Select patient
            if not self.system.patients:
                console.print("[bold red]Chưa có bệnh nhân trong hệ thống[/bold red]")
                return

            console.print("[bold]Danh sách bệnh nhân:[/bold]")
            patients_list = list(self.system.patients.values())
            for i, p in enumerate(patients_list, 1):
                console.print(f"  {i}. {p.name} ({p.id})")

            patient_choice = console.input("[bold]Chọn bệnh nhân (số):[/bold] ")
            patient_idx = int(patient_choice) - 1
            if patient_idx < 0 or patient_idx >= len(patients_list):
                raise ValueError("Bệnh nhân không hợp lệ")
            patient = patients_list[patient_idx]

            # Select department
            console.print("\n[bold]Danh sách khoa:[/bold]")
            depts_list = list(self.system.departments.keys())
            for i, dept_name in enumerate(depts_list, 1):
                console.print(f"  {i}. {dept_name}")

            dept_choice = console.input("[bold]Chọn khoa (số):[/bold] ")
            dept_idx = int(dept_choice) - 1
            if dept_idx < 0 or dept_idx >= len(depts_list):
                raise ValueError("Khoa không hợp lệ")
            department = self.system.departments[depts_list[dept_idx]]

            # Select doctor
            if not department.doctors:
                console.print("[bold red]Khoa này chưa có bác sĩ[/bold red]")
                return

            console.print("\n[bold]Danh sách bác sĩ:[/bold]")
            for i, doctor in enumerate(department.doctors, 1):
                console.print(f"  {i}. {doctor.name} ({doctor.id})")

            doctor_choice = console.input("[bold]Chọn bác sĩ (số):[/bold] ")
            doctor_idx = int(doctor_choice) - 1
            if doctor_idx < 0 or doctor_idx >= len(department.doctors):
                raise ValueError("Bác sĩ không hợp lệ")
            doctor = department.doctors[doctor_idx]

            # Time slot
            time_slot = console.input("[bold]Khung giờ (HH:MM, ví dụ 09:30):[/bold] ")
            if not Validation.validate_time_slot(time_slot):
                raise ValueError("Khung giờ không hợp lệ (định dạng HH:MM)")

            # Check slot availability (max 4 per hour)
            try:
                hour, minute = map(int, time_slot.split(":"))
                appointment_time = self.system.clock.get_time().replace(
                    hour=hour, minute=minute, second=0, microsecond=0
                )
            except ValueError:
                raise ValueError("Khung giờ không hợp lệ")

            if not doctor.is_available(appointment_time):
                console.print(
                    "[bold red]Bác sĩ đã kín lịch trong khung giờ này (tối đa 4 ca/giờ)[/bold red]"
                )
                return

            # Book the appointment
            success = doctor.book_appointment(patient, appointment_time)
            if success:
                patient.doctor_id = doctor.id
                console.print(f"\n[bold green]Đặt lịch thành công![/bold green]")
                console.print(f"Bệnh nhân: {patient.name}")
                console.print(f"Bác sĩ: {doctor.name}")
                console.print(
                    f"Thời gian: {appointment_time.strftime('%d/%m/%Y %H:%M')}"
                )
            else:
                console.print("[bold red]Đặt lịch thất bại[/bold red]")

        except ValueError as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")

    def view_appointments_by_doctor(self):
        """View appointments for a specific doctor."""
        all_doctors = []
        for dept in self.system.departments.values():
            all_doctors.extend(dept.doctors)

        if not all_doctors:
            console.print("[bold yellow]Chưa có bác sĩ nào[/bold yellow]")
            return

        console.print("[bold]Chọn bác sĩ:[/bold]")
        for i, doctor in enumerate(all_doctors, 1):
            console.print(f"  {i}. {doctor.name} ({doctor.id})")

        try:
            choice = int(console.input("[bold]Chọn (số):[/bold] ")) - 1
            if choice < 0 or choice >= len(all_doctors):
                raise ValueError()
            doctor = all_doctors[choice]
        except (ValueError, IndexError):
            console.print("[bold red]Bác sĩ không hợp lệ[/bold red]")
            return

        if not doctor.appointments:
            console.print(
                f"[bold yellow]{doctor.name} chưa có lịch hẹn nào[/bold yellow]"
            )
            return

        table = Table(
            title=f"LỊCH HẸN - {doctor.name}",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("Thời gian", width=20)
        table.add_column("Mã BN", width=10)
        table.add_column("Tên BN", width=25)
        table.add_column("Ưu tiên", width=12)

        for time_slot, patients in sorted(doctor.appointments.items()):
            time_str = time_slot.strftime("%d/%m/%Y %H:%M")
            for patient in patients:
                priority = self.PRIORITY_NAMES.get(
                    patient.priority, str(patient.priority)
                )
                table.add_row(time_str, patient.id, patient.name, priority)

        console.print(table)

    def view_appointments_by_date(self):
        """View all appointments for a specific date."""
        date_str = console.input("[bold]Nhập ngày (DD/MM/YYYY):[/bold] ")
        date = Validation.validate_date(date_str)
        if not date:
            console.print("[bold red]Ngày không hợp lệ[/bold red]")
            return

        all_appointments = []
        for dept in self.system.departments.values():
            for doctor in dept.doctors:
                for time_slot, patients in doctor.appointments.items():
                    if time_slot.date() == date.date():
                        for patient in patients:
                            all_appointments.append(
                                {
                                    "time": time_slot,
                                    "doctor": doctor,
                                    "patient": patient,
                                }
                            )

        if not all_appointments:
            console.print(
                f"[bold yellow]Không có lịch hẹn nào ngày {date_str}[/bold yellow]"
            )
            return

        table = Table(
            title=f"LỊCH HẸN NGÀY {date_str}",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("Giờ", width=10)
        table.add_column("Bác sĩ", width=20)
        table.add_column("Khoa", width=15)
        table.add_column("Mã BN", width=10)
        table.add_column("Tên BN", width=25)

        for app in sorted(all_appointments, key=lambda x: x["time"]):
            table.add_row(
                app["time"].strftime("%H:%M"),
                app["doctor"].name,
                app["doctor"].department,
                app["patient"].id,
                app["patient"].name,
            )

        console.print(table)

    def cancel_appointment(self):
        """Cancel an appointment."""
        all_doctors = []
        for dept in self.system.departments.values():
            all_doctors.extend(dept.doctors)

        if not all_doctors:
            console.print("[bold yellow]Chưa có bác sĩ nào[/bold yellow]")
            return

        console.print("[bold]Chọn bác sĩ:[/bold]")
        for i, doctor in enumerate(all_doctors, 1):
            console.print(f"  {i}. {doctor.name} ({doctor.id})")

        try:
            choice = int(console.input("[bold]Chọn (số):[/bold] ")) - 1
            if choice < 0 or choice >= len(all_doctors):
                raise ValueError()
            doctor = all_doctors[choice]
        except (ValueError, IndexError):
            console.print("[bold red]Bác sĩ không hợp lệ[/bold red]")
            return

        if not doctor.appointments:
            console.print(
                f"[bold yellow]{doctor.name} chưa có lịch hẹn nào[/bold yellow]"
            )
            return

        # List appointments
        appointments_list = []
        table = Table(
            title=f"LỊCH HẸN - {doctor.name}",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("STT", width=5)
        table.add_column("Thời gian", width=20)
        table.add_column("Mã BN", width=10)
        table.add_column("Tên BN", width=25)

        idx = 1
        for time_slot, patients in sorted(doctor.appointments.items()):
            for patient in patients:
                appointments_list.append((time_slot, patient))
                table.add_row(
                    str(idx),
                    time_slot.strftime("%d/%m/%Y %H:%M"),
                    patient.id,
                    patient.name,
                )
                idx += 1

        console.print(table)

        try:
            cancel_idx = (
                int(console.input("[bold]Chọn lịch hẹn cần hủy (STT):[/bold] ")) - 1
            )
            if cancel_idx < 0 or cancel_idx >= len(appointments_list):
                raise ValueError()
            time_slot, patient = appointments_list[cancel_idx]
        except (ValueError, IndexError):
            console.print("[bold red]Lịch hẹn không hợp lệ[/bold red]")
            return

        confirm = Confirm.ask(
            f"Hủy lịch hẹn của {patient.name} lúc {time_slot.strftime('%H:%M')}?",
            default=False,
        )

        if confirm:
            if time_slot in doctor.appointments:
                doctor.appointments[time_slot] = [
                    p for p in doctor.appointments[time_slot] if p.id != patient.id
                ]
                if not doctor.appointments[time_slot]:
                    del doctor.appointments[time_slot]
            patient.appointment_time = None
            console.print("[bold green]Đã hủy lịch hẹn thành công[/bold green]")

    # ═══════════════════════════════════════════════════════════════
    # CHECK-IN
    # ═══════════════════════════════════════════════════════════════

    def checkin_menu(self):
        """Check-in sub-menu."""
        while True:
            self.clear_screen()
            console.print(
                Panel("[bold]CHECK-IN BỆNH NHÂN[/bold]", border_style="yellow")
            )
            console.print("  [1] Check-in bệnh nhân")
            console.print("  [2] Xem hàng đợi hiện tại")
            console.print("  [3] Xem bệnh nhân cấp cứu")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 3)

            if choice == 1:
                self.checkin_patient()
            elif choice == 2:
                self.view_queues()
            elif choice == 3:
                self.view_emergency_patients()
            elif choice == 0:
                break

            self.pause()

    def checkin_patient(self):
        """Check in a patient to a department."""
        console.print(Panel("[bold yellow]CHECK-IN BỆNH NHÂN[/bold yellow]"))

        try:
            # Select patient
            if not self.system.patients:
                console.print("[bold red]Chưa có bệnh nhân trong hệ thống[/bold red]")
                return

            # Show waiting patients first
            waiting_patients = [
                p for p in self.system.patients.values() if p.status == "waiting"
            ]
            if waiting_patients:
                console.print("[bold]Bệnh nhân đang chờ:[/bold]")
                for i, p in enumerate(waiting_patients, 1):
                    console.print(f"  {i}. {p.name} ({p.id})")

            console.print("\n[bold]Tất cả bệnh nhân:[/bold]")
            patients_list = list(self.system.patients.values())
            for i, p in enumerate(patients_list, 1):
                status = self.STATUS_COLORS.get(p.status, p.status)
                console.print(f"  {i}. {p.name} ({p.id}) - {status}")

            patient_choice = console.input("[bold]Chọn bệnh nhân (số):[/bold] ")
            patient_idx = int(patient_choice) - 1
            if patient_idx < 0 or patient_idx >= len(patients_list):
                raise ValueError("Bệnh nhân không hợp lệ")
            patient = patients_list[patient_idx]

            # Select department
            console.print("\n[bold]Chọn khoa:[/bold]")
            depts_list = list(self.system.departments.keys())
            for i, dept_name in enumerate(depts_list, 1):
                console.print(f"  {i}. {dept_name}")

            dept_choice = console.input("[bold]Chọn khoa (số):[/bold] ")
            dept_idx = int(dept_choice) - 1
            if dept_idx < 0 or dept_idx >= len(depts_list):
                raise ValueError("Khoa không hợp lệ")
            department = depts_list[dept_idx]

            # Priority confirmation
            current_priority = self.PRIORITY_NAMES.get(
                patient.priority, str(patient.priority)
            )
            console.print(f"\n[bold]Mức độ ưu tiên hiện tại:[/bold] {current_priority}")
            change_priority = Confirm.ask("Thay đổi mức độ ưu tiên?", default=False)

            if change_priority:
                console.print("\n[bold]Chọn mức độ ưu tiên mới:[/bold]")
                console.print("  1 - Cấp cứu")
                console.print("  2 - Ưu tiên")
                console.print("  3 - Bình thường")
                console.print("  4 - Thấp")
                priority_str = console.input("[bold]Chọn (1-4):[/bold] ")
                patient.priority = Validation.validate_priority(priority_str)

            # Check in
            self.system.check_in_patient(patient.id, department)

            console.print(f"\n[bold green]Check-in thành công![/bold green]")
            console.print(f"Bệnh nhân: {patient.name}")
            console.print(f"Khoa: {department}")
            console.print(
                f"Ưu tiên: {self.PRIORITY_NAMES.get(patient.priority, str(patient.priority))}"
            )
            console.print(f"Thời gian: {self.system.clock.get_time_str()}")

            # Check if there are available doctors
            dept_obj = self.system.departments.get(department)
            if dept_obj:
                available = len(dept_obj.get_available_doctors())
                if available > 0:
                    console.print(
                        f"[green]Có {available} bác sĩ sẵn sàng tiếp nhận[/green]"
                    )
                else:
                    console.print(
                        "[yellow]Hiện tại chưa có bác sĩ rảnh, vui lòng chờ[/yellow]"
                    )

        except ValueError as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Lỗi: {e}[/bold red]")

    def view_queues(self):
        """View all department queues."""
        if not self.system.departments:
            console.print("[bold yellow]Chưa có khoa nào[/bold yellow]")
            return

        for dept_name, dept in self.system.departments.items():
            if not dept.queue:
                continue

            table = Table(
                title=f"HÀNG ĐỢI - {dept_name}",
                box=box.SIMPLE,
                header_style="bold cyan",
            )
            table.add_column("STT", width=5)
            table.add_column("Mã BN", width=10)
            table.add_column("Tên BN", width=25)
            table.add_column("Ưu tiên", width=12)
            table.add_column("Trạng thái", width=14)

            for i, patient in enumerate(dept.queue, 1):
                priority = self.PRIORITY_NAMES.get(
                    patient.priority, str(patient.priority)
                )
                status = self.STATUS_COLORS.get(patient.status, patient.status)
                table.add_row(str(i), patient.id, patient.name, priority, status)

            console.print(table)
            console.print()

        # Emergency patients
        if self.system.emergency_patients:
            table = Table(
                title="[bold red]CẤP CỨU[/bold red]",
                box=box.SIMPLE,
                header_style="bold red",
            )
            table.add_column("STT", width=5)
            table.add_column("Mã BN", width=10)
            table.add_column("Tên BN", width=25)
            table.add_column("Ưu tiên", width=12)

            for i, patient in enumerate(self.system.emergency_patients, 1):
                priority = self.PRIORITY_NAMES.get(
                    patient.priority, str(patient.priority)
                )
                table.add_row(str(i), patient.id, patient.name, priority)

            console.print(table)

    def view_emergency_patients(self):
        """View emergency patients."""
        if not self.system.emergency_patients:
            console.print("[bold yellow]Không có bệnh nhân cấp cứu[/bold yellow]")
            return

        table = Table(
            title="[bold red]BỆNH NHÂN CẤP CỨU[/bold red]",
            box=box.ROUNDED,
            header_style="bold red",
        )
        table.add_column("STT", width=5)
        table.add_column("Mã BN", width=10)
        table.add_column("Tên BN", width=25)
        table.add_column("Check-in", width=20)

        for i, patient in enumerate(self.system.emergency_patients, 1):
            checkin_time = (
                patient.checked_in_at.strftime("%d/%m %H:%M")
                if patient.checked_in_at
                else "N/A"
            )
            table.add_row(str(i), patient.id, patient.name, checkin_time)

        console.print(table)

    # ═══════════════════════════════════════════════════════════════
    # DOCTOR EXAMINATION
    # ═══════════════════════════════════════════════════════════════

    def examination_menu(self):
        """Doctor examination sub-menu."""
        while True:
            self.clear_screen()
            console.print(Panel("[bold]BÁC SĨ KHÁM BỆNH[/bold]", border_style="yellow"))
            console.print("  [1] Tiếp nhận bệnh nhân tiếp theo")
            console.print("  [2] Hoàn thành khám bệnh")
            console.print("  [3] Xem bệnh nhân đang khám")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 3)

            if choice == 1:
                self.start_examination()
            elif choice == 2:
                self.complete_examination()
            elif choice == 3:
                self.view_current_examinations()
            elif choice == 0:
                break

            self.pause()

    def start_examination(self):
        """Assign next patient to a doctor."""
        console.print(Panel("[bold yellow]TIẾP NHẬN BỆNH NHÂN[/bold yellow]"))

        # Select department
        console.print("[bold]Chọn khoa:[/bold]")
        depts_list = list(self.system.departments.keys())
        for i, dept_name in enumerate(depts_list, 1):
            dept = self.system.departments[dept_name]
            available = len(dept.get_available_doctors())
            waiting = dept.get_queue_length()
            console.print(
                f"  {i}. {dept_name} (BS rảnh: {available}, BN chờ: {waiting})"
            )

        try:
            dept_choice = int(console.input("[bold]Chọn khoa (số):[/bold] ")) - 1
            if dept_choice < 0 or dept_choice >= len(depts_list):
                raise ValueError()
            dept_name = depts_list[dept_choice]
        except (ValueError, IndexError):
            console.print("[bold red]Khoa không hợp lệ[/bold red]")
            return

        patient = self.system.assign_patient_to_doctor(dept_name)

        if patient:
            # Find the doctor
            doctor = None
            for d in self.system.departments[dept_name].doctors:
                if d.current_patient and d.current_patient.id == patient.id:
                    doctor = d
                    break

            console.print(f"\n[bold green]Đã tiếp nhận bệnh nhân![/bold green]")
            console.print(f"Bệnh nhân: {patient.name} ({patient.id})")
            if doctor:
                console.print(f"Bác sĩ: {doctor.name} ({doctor.id})")
            console.print(
                f"Ưu tiên: {self.PRIORITY_NAMES.get(patient.priority, str(patient.priority))}"
            )

            # Advance clock
            self.system.clock.tick(15)
            console.print(f"[dim]Thời gian: {self.system.clock.get_time_str()}[/dim]")
        else:
            console.print(
                "[bold yellow]Không thể tiếp nhận: Không có bác sĩ rảnh hoặc không có bệnh nhân chờ[/bold yellow]"
            )

    def complete_examination(self):
        """Complete an examination."""
        console.print(Panel("[bold yellow]HOÀN THÀNH KHÁM BỆNH[/bold yellow]"))

        # Find doctors with current patients
        busy_doctors = []
        for dept in self.system.departments.values():
            for doctor in dept.doctors:
                if doctor.current_patient:
                    busy_doctors.append(doctor)

        if not busy_doctors:
            console.print(
                "[bold yellow]Không có bác sĩ nào đang khám bệnh[/bold yellow]"
            )
            return

        console.print("[bold]Chọn bác sĩ:[/bold]")
        for i, doctor in enumerate(busy_doctors, 1):
            console.print(
                f"  {i}. {doctor.name} - {doctor.current_patient.name} ({doctor.current_patient.id})"
            )

        try:
            choice = int(console.input("[bold]Chọn (số):[/bold] ")) - 1
            if choice < 0 or choice >= len(busy_doctors):
                raise ValueError()
            doctor = busy_doctors[choice]
        except (ValueError, IndexError):
            console.print("[bold red]Bác sĩ không hợp lệ[/bold red]")
            return

        patient = doctor.current_patient

        console.print(f"\n[bold]Hoàn thành khám cho {patient.name}[/bold]")

        # Services provided
        services = []
        service_options = list(self.billing_system.service_catalog.keys())

        while True:
            console.print("\n[bold]Dịch vụ đã cung cấp:[/bold]")
            for i, service in enumerate(service_options, 1):
                cost = self.billing_system.service_catalog[service]
                console.print(f"  {i}. {service} ({cost:,.0f} VNĐ)")
            console.print("  0. Kết thúc chọn dịch vụ")

            try:
                service_choice = int(console.input("[bold]Chọn dịch vụ (số):[/bold] "))
                if service_choice == 0:
                    break
                if service_choice < 1 or service_choice > len(service_options):
                    raise ValueError()
                services.append(service_options[service_choice - 1])
                console.print(
                    f"[green]Đã thêm: {service_options[service_choice - 1]}[/green]"
                )
            except (ValueError, IndexError):
                console.print("[bold red]Dịch vụ không hợp lệ[/bold red]")

        # Complete examination
        completed_patient = self.system.complete_examination(doctor.id)

        if completed_patient:
            # Generate bill
            if services:
                bill = self.billing_system.create_bill(
                    patient_id=patient.id,
                    doctor_id=doctor.id,
                    dept_id=doctor.department,
                    services=services,
                )
                console.print(
                    f"\n[bold green]Đã tạo hóa đơn: {bill.bill_id}[/bold green]"
                )
                console.print(f"Tổng tiền: [bold]{bill.total_amount:,.0f} VNĐ[/bold]")

            console.print(
                f"\n[bold green]Hoàn thành khám bệnh cho {completed_patient.name}[/bold green]"
            )
            self.system.clock.tick(15)
            console.print(f"[dim]Thời gian: {self.system.clock.get_time_str()}[/dim]")

    def view_current_examinations(self):
        """View all current examinations."""
        busy_doctors = []
        for dept in self.system.departments.values():
            for doctor in dept.doctors:
                if doctor.current_patient:
                    busy_doctors.append(doctor)

        if not busy_doctors:
            console.print("[bold yellow]Không có bác sĩ nào đang khám[/bold yellow]")
            return

        table = Table(
            title="BỆNH NHÂN ĐANG KHÁM",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("Bác sĩ", width=20)
        table.add_column("Khoa", width=15)
        table.add_column("Mã BN", width=10)
        table.add_column("Tên BN", width=25)
        table.add_column("Ưu tiên", width=12)

        for doctor in busy_doctors:
            patient = doctor.current_patient
            priority = self.PRIORITY_NAMES.get(patient.priority, str(patient.priority))
            table.add_row(
                doctor.name, doctor.department, patient.id, patient.name, priority
            )

        console.print(table)

    # ═══════════════════════════════════════════════════════════════
    # DASHBOARD
    # ═══════════════════════════════════════════════════════════════

    def dashboard_menu(self):
        """Display real-time dashboard."""
        self.clear_screen()
        console.print(
            Panel(
                f"[bold cyan]DASHBOARD - {self.system.clock.get_time().strftime('%d/%m/%Y %H:%M')}[/bold cyan]",
                border_style="cyan",
            )
        )

        # Statistics panels
        total_patients = len(self.system.patients)
        waiting_patients = len(
            [p for p in self.system.patients.values() if p.status == "waiting"]
        )
        in_exam = len(
            [p for p in self.system.patients.values() if p.status == "in_exam"]
        )
        completed = len(
            [p for p in self.system.patients.values() if p.status == "completed"]
        )
        emergency_count = len(self.system.emergency_patients)

        # Create stats layout
        stats_table = Table(box=box.SIMPLE, show_header=False)
        stats_table.add_column("Metric", style="bold")
        stats_table.add_column("Value", justify="right")
        stats_table.add_row("Tổng bệnh nhân", str(total_patients))
        stats_table.add_row("Đang chờ", f"[yellow]{waiting_patients}[/yellow]")
        stats_table.add_row("Đang khám", f"[blue]{in_exam}[/blue]")
        stats_table.add_row("Hoàn thành", f"[green]{completed}[/green]")
        stats_table.add_row("Cấp cứu", f"[bold red]{emergency_count}[/bold red]")

        # Department stats
        dept_table = Table(
            title="KHOA",
            box=box.SIMPLE,
            header_style="bold",
            show_header=True,
        )
        dept_table.add_column("Khoa", width=20)
        dept_table.add_column("BS", justify="center", width=6)
        dept_table.add_column("Chờ", justify="center", width=6)
        dept_table.add_column("Rảnh", justify="center", width=6)

        for dept in self.system.departments.values():
            available = len(dept.get_available_doctors())
            waiting = dept.get_queue_length()
            total_docs = len(dept.doctors)
            dept_table.add_row(
                dept.name,
                str(total_docs),
                f"[yellow]{waiting}[/yellow]" if waiting > 0 else str(waiting),
                f"[green]{available}[/green]" if available > 0 else str(available),
            )

        # Doctor status
        doctor_table = Table(
            title="BÁC SĨ",
            box=box.SIMPLE,
            header_style="bold",
        )
        doctor_table.add_column("Bác sĩ", width=20)
        doctor_table.add_column("Khoa", width=15)
        doctor_table.add_column("Trạng thái", justify="center", width=12)
        doctor_table.add_column("BN hiện tại", width=20)

        for dept in self.system.departments.values():
            for doctor in dept.doctors:
                if doctor.current_patient:
                    status = "[red]Bận[/red]"
                    current = doctor.current_patient.name
                else:
                    status = "[green]Rảnh[/green]"
                    current = "-"
                doctor_table.add_row(doctor.name, doctor.department, status, current)

        # Revenue
        total_revenue = self.billing_system.get_total_revenue()
        total_bills = len(self.billing_system.bills)
        paid_bills = len([b for b in self.billing_system.bills.values() if b.is_paid])

        revenue_table = Table(box=box.SIMPLE, show_header=False)
        revenue_table.add_column("Metric", style="bold")
        revenue_table.add_column("Value", justify="right")
        revenue_table.add_row("Tổng doanh thu", f"{total_revenue:,.0f} VNĐ")
        revenue_table.add_row("Tổng hóa đơn", str(total_bills))
        revenue_table.add_row("Đã thanh toán", f"[green]{paid_bills}[/green]")

        # Display all panels
        console.print(
            Panel(stats_table, title="[bold]THỐNG Kê[/bold]", border_style="green")
        )
        console.print()
        console.print(Panel(dept_table, title="[bold]KHOA[/bold]", border_style="blue"))
        console.print()
        console.print(
            Panel(doctor_table, title="[bold]BÁC SĨ[/bold]", border_style="yellow")
        )
        console.print()
        console.print(
            Panel(revenue_table, title="[bold]DOANH THU[/bold]", border_style="magenta")
        )

    # ═══════════════════════════════════════════════════════════════
    # BILLING
    # ═══════════════════════════════════════════════════════════════

    def billing_menu(self):
        """Billing sub-menu."""
        while True:
            self.clear_screen()
            console.print(
                Panel("[bold]TÍNH TIỀN / HÓA ĐƠN[/bold]", border_style="magenta")
            )
            console.print("  [1] Xem danh sách hóa đơn")
            console.print("  [2] Xem hóa đơn theo bệnh nhân")
            console.print("  [3] Xem chi tiết hóa đơn")
            console.print("  [4] Thanh toán hóa đơn")
            console.print("  [5] Thêm dịch vụ vào hóa đơn")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 5)

            if choice == 1:
                self.view_all_bills()
            elif choice == 2:
                self.view_patient_bills()
            elif choice == 3:
                self.view_bill_detail()
            elif choice == 4:
                self.pay_bill()
            elif choice == 5:
                self.add_service_to_bill()
            elif choice == 0:
                break

            self.pause()

    def view_all_bills(self):
        """Display all bills."""
        if not self.billing_system.bills:
            console.print("[bold yellow]Chưa có hóa đơn nào[/bold yellow]")
            return

        table = Table(
            title="DANH SÁCH HÓA ĐƠN",
            box=box.ROUNDED,
            header_style="bold magenta",
        )
        table.add_column("Mã HD", style="bold", width=12)
        table.add_column("Mã BN", width=10)
        table.add_column("Mã BS", width=10)
        table.add_column("Khoa", width=15)
        table.add_column("Tổng tiền", justify="right", width=15)
        table.add_column("Trạng thái", justify="center", width=12)

        for bill in self.billing_system.bills.values():
            status = "[green]ĐÃ TT[/green]" if bill.is_paid else "[red]CHƯA TT[/red]"
            table.add_row(
                bill.bill_id,
                bill.patient_id,
                bill.doctor_id,
                bill.department_id,
                f"{bill.total_amount:,.0f}",
                status,
            )

        console.print(table)
        console.print(f"\n[dim]Tổng số: {len(self.billing_system.bills)} hóa đơn[/dim]")

    def view_patient_bills(self):
        """View bills for a specific patient."""
        patient_id = console.input("[bold]Nhập mã bệnh nhân:[/bold] ")

        if patient_id not in self.system.patients:
            console.print(f"[bold red]Không tìm thấy bệnh nhân {patient_id}[/bold red]")
            return

        bills = self.billing_system.get_patient_bills(patient_id)
        if not bills:
            console.print(
                f"[bold yellow]Bệnh nhân {patient_id} chưa có hóa đơn nào[/bold yellow]"
            )
            return

        patient = self.system.patients[patient_id]
        table = Table(
            title=f"HÓA ĐƠN - {patient.name}",
            box=box.ROUNDED,
            header_style="bold magenta",
        )
        table.add_column("Mã HD", width=12)
        table.add_column("Ngày", width=18)
        table.add_column("Tổng tiền", justify="right", width=15)
        table.add_column("Trạng thái", justify="center", width=12)

        for bill in bills:
            status = "[green]ĐÃ TT[/green]" if bill.is_paid else "[red]CHƯA TT[/red]"
            date_str = bill.created_at.strftime("%d/%m/%Y %H:%M")
            table.add_row(bill.bill_id, date_str, f"{bill.total_amount:,.0f}", status)

        console.print(table)

    def view_bill_detail(self):
        """View detailed bill information."""
        bill_id = console.input("[bold]Nhập mã hóa đơn:[/bold] ")

        if bill_id not in self.billing_system.bills:
            console.print(f"[bold red]Không tìm thấy hóa đơn {bill_id}[/bold red]")
            return

        bill = self.billing_system.bills[bill_id]
        patient = self.system.patients.get(bill.patient_id)

        info = []
        info.append(f"[bold]Mã hóa đơn:[/bold] {bill.bill_id}")
        info.append(
            f"[bold]Bệnh nhân:[/bold] {patient.name if patient else 'N/A'} ({bill.patient_id})"
        )
        info.append(f"[bold]Bác sĩ:[/bold] {bill.doctor_id}")
        info.append(f"[bold]Khoa:[/bold] {bill.department_id}")
        info.append(
            f"[bold]Ngày tạo:[/bold] {bill.created_at.strftime('%d/%m/%Y %H:%M')}"
        )

        info.append("\n[bold]Chi tiết dịch vụ:[/bold]")
        for service, cost in bill.items:
            info.append(f"  - {service}: {cost:,.0f} VNĐ")

        info.append(f"\n[bold]Tổng cộng:[/bold] {bill.total_amount:,.0f} VNĐ")
        status = "Đã thanh toán" if bill.is_paid else "Chưa thanh toán"
        info.append(f"[bold]Trạng thái:[/bold] {status}")

        panel = Panel(
            "\n".join(info),
            title=f"[bold]HÓA ĐƠN {bill_id}[/bold]",
            border_style="magenta",
            box=box.ROUNDED,
        )
        console.print(panel)

    def pay_bill(self):
        """Mark a bill as paid."""
        bill_id = console.input("[bold]Nhập mã hóa đơn cần thanh toán:[/bold] ")

        if bill_id not in self.billing_system.bills:
            console.print(f"[bold red]Không tìm thấy hóa đơn {bill_id}[/bold red]")
            return

        bill = self.billing_system.bills[bill_id]
        if bill.is_paid:
            console.print("[bold yellow]Hóa đơn này đã được thanh toán[/bold yellow]")
            return

        console.print(f"\n[bold]Tổng tiền:[/bold] {bill.total_amount:,.0f} VNĐ")
        confirm = Confirm.ask("Xác nhận thanh toán?", default=False)

        if confirm:
            self.billing_system.mark_bill_paid(bill_id)
            console.print("[bold green]Thanh toán thành công![/bold green]")
        else:
            console.print("[dim]Đã hủy thanh toán[/dim]")

    def add_service_to_bill(self):
        """Add a service to an existing bill."""
        bill_id = console.input("[bold]Nhập mã hóa đơn:[/bold] ")

        if bill_id not in self.billing_system.bills:
            console.print(f"[bold red]Không tìm thấy hóa đơn {bill_id}[/bold red]")
            return

        console.print("[bold]Danh sách dịch vụ:[/bold]")
        services = list(self.billing_system.service_catalog.keys())
        for i, service in enumerate(services, 1):
            cost = self.billing_system.service_catalog[service]
            console.print(f"  {i}. {service} ({cost:,.0f} VNĐ)")

        try:
            choice = int(console.input("[bold]Chọn dịch vụ (số):[/bold] ")) - 1
            if choice < 0 or choice >= len(services):
                raise ValueError()
            service_name = services[choice]
        except (ValueError, IndexError):
            console.print("[bold red]Dịch vụ không hợp lệ[/bold red]")
            return

        success = self.billing_system.add_service_to_bill(bill_id, service_name)
        if success:
            bill = self.billing_system.bills[bill_id]
            console.print(f"[bold green]Đã thêm dịch vụ![/bold green]")
            console.print(f"Tổng tiền mới: {bill.total_amount:,.0f} VNĐ")
        else:
            console.print("[bold red]Không thể thêm dịch vụ[/bold red]")

    # ═══════════════════════════════════════════════════════════════
    # REPORTS & STATISTICS
    # ═══════════════════════════════════════════════════════════════

    def reports_menu(self):
        """Reports and statistics sub-menu."""
        while True:
            self.clear_screen()
            console.print(
                Panel("[bold]BÁO CÁO & THỐNG KÊ[/bold]", border_style="magenta")
            )
            console.print("  [1] Thống kê tổng quan")
            console.print("  [2] Thống kê theo khoa")
            console.print("  [3] Xuất dữ liệu ra JSON")
            console.print("  [4] Xuất dữ liệu ra CSV")
            console.print("  [5] Xuất dữ liệu ra Excel")
            console.print("  [0] Quay lại menu chính")
            console.print()

            choice = self.get_menu_choice("Chọn chức năng", 5)

            if choice == 1:
                self.general_statistics()
            elif choice == 2:
                self.department_statistics()
            elif choice == 3:
                self.export_json()
            elif choice == 4:
                self.export_csv()
            elif choice == 5:
                self.export_excel()
            elif choice == 0:
                break

            self.pause()

    def general_statistics(self):
        """Display general statistics."""
        total_patients = len(self.system.patients)
        total_doctors = sum(len(d.doctors) for d in self.system.departments.values())
        total_depts = len(self.system.departments)

        waiting = len(
            [p for p in self.system.patients.values() if p.status == "waiting"]
        )
        in_exam = len(
            [p for p in self.system.patients.values() if p.status == "in_exam"]
        )
        completed = len(
            [p for p in self.system.patients.values() if p.status == "completed"]
        )
        billed = len([p for p in self.system.patients.values() if p.status == "billed"])
        emergency = len(self.system.emergency_patients)

        total_revenue = self.billing_system.get_total_revenue()
        total_bills = len(self.billing_system.bills)
        paid_bills = len([b for b in self.billing_system.bills.values() if b.is_paid])
        unpaid_bills = total_bills - paid_bills

        table = Table(
            title="THỐNG KÊ TỔNG QUAN",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("Chỉ số", style="bold", width=30)
        table.add_column("Giá trị", justify="right", width=20)

        table.add_row("Tổng bệnh nhân", str(total_patients))
        table.add_row("Tổng bác sĩ", str(total_doctors))
        table.add_row("Tổng khoa", str(total_depts))
        table.add_section()
        table.add_row("Đang chờ", f"[yellow]{waiting}[/yellow]")
        table.add_row("Đang khám", f"[blue]{in_exam}[/blue]")
        table.add_row("Hoàn thành", f"[green]{completed}[/green]")
        table.add_row("Đã tính tiền", f"[cyan]{billed}[/cyan]")
        table.add_row("Cấp cứu", f"[bold red]{emergency}[/bold red]")
        table.add_section()
        table.add_row("Tổng hóa đơn", str(total_bills))
        table.add_row("Đã thanh toán", f"[green]{paid_bills}[/green]")
        table.add_row("Chưa thanh toán", f"[red]{unpaid_bills}[/red]")
        table.add_row("Tổng doanh thu", f"[bold]{total_revenue:,.0f} VNĐ[/bold]")

        console.print(table)

    def department_statistics(self):
        """Display statistics by department."""
        table = Table(
            title="THỐNG KÊ THEO KHOA",
            box=box.ROUNDED,
            header_style="bold cyan",
        )
        table.add_column("Khoa", style="bold", width=20)
        table.add_column("BS", justify="center", width=6)
        table.add_column("BN chờ", justify="center", width=8)
        table.add_column("BS rảnh", justify="center", width=8)
        table.add_column("Doanh thu", justify="right", width=15)

        for dept_name, dept in self.system.departments.items():
            revenue = self.billing_system.get_department_revenue(dept_name)
            available = len(dept.get_available_doctors())
            table.add_row(
                dept_name,
                str(len(dept.doctors)),
                str(dept.get_queue_length()),
                str(available),
                f"{revenue:,.0f}",
            )

        console.print(table)

    def export_json(self):
        """Export data to JSON files."""
        export_dir = Path("exports")
        export_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            # Export patients
            patients_data = {}
            for pid, patient in self.system.patients.items():
                patients_data[pid] = {
                    "id": patient.id,
                    "name": patient.name,
                    "priority": patient.priority,
                    "status": patient.status,
                    "department": patient.department,
                }

            patients_file = export_dir / f"patients_{timestamp}.json"
            patients_file.write_text(
                json.dumps(patients_data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            # Export bills
            bills_data = {}
            for bill_id, bill in self.billing_system.bills.items():
                bills_data[bill_id] = {
                    "bill_id": bill.bill_id,
                    "patient_id": bill.patient_id,
                    "total_amount": bill.total_amount,
                    "is_paid": bill.is_paid,
                    "created_at": bill.created_at.isoformat(),
                }

            bills_file = export_dir / f"bills_{timestamp}.json"
            bills_file.write_text(
                json.dumps(bills_data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            console.print(
                f"[bold green]Đã xuất dữ liệu JSON vào thư mục {export_dir}[/bold green]"
            )
            console.print(f"  - {patients_file.name}")
            console.print(f"  - {bills_file.name}")

        except Exception as e:
            console.print(f"[bold red]Lỗi khi xuất JSON: {e}[/bold red]")

    def export_csv(self):
        """Export data to CSV files."""
        export_dir = Path("exports")
        export_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            # Export patients
            patients_file = export_dir / f"patients_{timestamp}.csv"
            with open(patients_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(
                    ["Mã BN", "Họ tên", "Tuổi", "Ưu tiên", "Trạng thái", "Khoa"]
                )
                for patient in self.system.patients.values():
                    writer.writerow(
                        [
                            patient.id,
                            patient.name,
                            getattr(patient, "age", ""),
                            patient.priority,
                            patient.status,
                            patient.department or "",
                        ]
                    )

            # Export bills
            bills_file = export_dir / f"bills_{timestamp}.csv"
            with open(bills_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(
                    [
                        "Mã HD",
                        "Mã BN",
                        "Mã BS",
                        "Khoa",
                        "Tổng tiền",
                        "Trạng thái",
                        "Ngày tạo",
                    ]
                )
                for bill in self.billing_system.bills.values():
                    writer.writerow(
                        [
                            bill.bill_id,
                            bill.patient_id,
                            bill.doctor_id,
                            bill.department_id,
                            bill.total_amount,
                            "Đã TT" if bill.is_paid else "Chưa TT",
                            bill.created_at.strftime("%d/%m/%Y %H:%M"),
                        ]
                    )

            console.print(
                f"[bold green]Đã xuất dữ liệu CSV vào thư mục {export_dir}[/bold green]"
            )
            console.print(f"  - {patients_file.name}")
            console.print(f"  - {bills_file.name}")

        except Exception as e:
            console.print(f"[bold red]Lỗi khi xuất CSV: {e}[/bold red]")

    def export_excel(self):
        """Export data to Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            console.print(
                "[bold yellow]Thiếu thư viện openpyxl. Cài đặt: pip install openpyxl[/bold yellow]"
            )
            return

        export_dir = Path("exports")
        export_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = export_dir / f"hospital_data_{timestamp}.xlsx"

        try:
            wb = openpyxl.Workbook()

            # Patients sheet
            ws_patients = wb.active
            ws_patients.title = "Patients"
            headers = ["Mã BN", "Họ tên", "Tuổi", "Ưu tiên", "Trạng thái", "Khoa"]
            ws_patients.append(headers)
            for h in ws_patients[1]:
                h.font = Font(bold=True)
                h.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )

            for patient in self.system.patients.values():
                ws_patients.append(
                    [
                        patient.id,
                        patient.name,
                        getattr(patient, "age", ""),
                        patient.priority,
                        patient.status,
                        patient.department or "",
                    ]
                )

            # Bills sheet
            ws_bills = wb.create_sheet(title="Bills")
            headers = [
                "Mã HD",
                "Mã BN",
                "Mã BS",
                "Khoa",
                "Tổng tiền",
                "Trạng thái",
                "Ngày tạo",
            ]
            ws_bills.append(headers)
            for h in ws_bills[1]:
                h.font = Font(bold=True)
                h.fill = PatternFill(
                    start_color="70AD47", end_color="70AD47", fill_type="solid"
                )

            for bill in self.billing_system.bills.values():
                ws_bills.append(
                    [
                        bill.bill_id,
                        bill.patient_id,
                        bill.doctor_id,
                        bill.department_id,
                        bill.total_amount,
                        "Đã TT" if bill.is_paid else "Chưa TT",
                        bill.created_at.strftime("%d/%m/%Y %H:%M"),
                    ]
                )

            wb.save(excel_file)
            console.print(
                f"[bold green]Đã xuất dữ liệu Excel: {excel_file}[/bold green]"
            )

        except Exception as e:
            console.print(f"[bold red]Lỗi khi xuất Excel: {e}[/bold red]")

    # ═══════════════════════════════════════════════════════════════
    # SAMPLE DATA
    # ═══════════════════════════════════════════════════════════════

    def sample_data_menu(self):
        """Generate sample data."""
        self.clear_screen()
        console.print(Panel("[bold]SINH DỮ LIỆU MẪU[/bold]", border_style="white"))

        count_str = (
            console.input("[bold]Số lượng bệnh nhân mẫu (mặc định 10):[/bold] ") or "10"
        )
        try:
            count = int(count_str)
            if count < 1 or count > 100:
                raise ValueError("Số lượng phải từ 1 đến 100")
        except ValueError:
            console.print("[bold red]Số lượng không hợp lệ[/bold red]")
            return

        import random

        vietnamese_names = [
            "Nguyễn Văn A",
            "Trần Thị B",
            "Lê Văn C",
            "Phạm Thị D",
            "Hoàng Văn E",
            "Ngô Thị F",
            "Đỗ Văn G",
            "Bùi Thị H",
            "Dương Văn I",
            "Lý Thị K",
            "Vũ Văn L",
            "Nguyễn Thị M",
            "Trần Văn N",
            "Lê Thị O",
            "Phạm Văn P",
            "Hoàng Thị Q",
            "Đỗ Văn R",
            "Bùi Thị S",
            "Dương Văn T",
            "Lý Thị U",
        ]

        symptoms_list = [
            "Đau đầu, sốt",
            "Đau bụng, buồn nôn",
            "Khó thở, ho",
            "Đau ngực, mệt mỏi",
            "Đau lưng, tê chân",
            "Chấn thương, chảy máu",
            "Tiêu chảy, đau bụng",
            "Dị ứng, phát ban",
            "Mờ mắt, đau mắt",
            "Đau tai, chảy máu tai",
        ]

        depts_list = list(self.system.departments.keys())

        generated = 0
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            console=console,
        ) as progress:
            task = progress.add_task("Đang tạo dữ liệu...", total=count)

            for i in range(count):
                self.patient_counter += 1
                patient_id = f"BN{self.patient_counter:04d}"

                name = random.choice(vietnamese_names)
                if self.system.patients:
                    # Make name unique if needed
                    suffix = 1
                    base_name = name
                    while any(p.name == name for p in self.system.patients.values()):
                        name = f"{base_name} {suffix}"
                        suffix += 1

                age = random.randint(18, 80)
                priority = random.choices([1, 2, 3, 4], weights=[5, 15, 60, 20])[0]
                symptoms = random.choice(symptoms_list)
                phone = f"09{random.randint(10000000, 99999999)}"
                address = f"Số {random.randint(1, 200)}, Đường ABC, Quận {random.randint(1, 12)}, TP.HCM"

                patient = Patient(id=patient_id, name=name, priority=priority)
                patient.age = age
                patient.symptoms = symptoms
                patient.phone = phone
                patient.address = address

                self.system.register_patient(patient)

                # Auto check-in some patients
                if random.random() < 0.7 and depts_list:
                    dept = random.choice(depts_list)
                    try:
                        self.system.check_in_patient(patient_id, dept)
                    except:
                        pass

                generated += 1
                progress.update(task, advance=1)

        console.print(f"\n[bold green]Đã tạo {generated} bệnh nhân mẫu[/bold green]")

    # ═══════════════════════════════════════════════════════════════
    # MAIN LOOP
    # ═══════════════════════════════════════════════════════════════

    def run(self):
        """Main application loop."""
        console.print(
            Panel(
                "[bold green]Chào mừng đến với Hệ thống Xếp hạng Khám bệnh[/bold green]",
                border_style="green",
            )
        )

        while self.running:
            self.display_main_menu()
            choice = self.get_menu_choice("Chọn chức năng", 10)

            if choice == 1:
                self.patient_menu()
            elif choice == 2:
                self.doctor_menu()
            elif choice == 3:
                self.department_menu()
            elif choice == 4:
                self.appointment_menu()
            elif choice == 5:
                self.checkin_menu()
            elif choice == 6:
                self.examination_menu()
            elif choice == 7:
                self.dashboard_menu()
                self.pause()
            elif choice == 8:
                self.billing_menu()
            elif choice == 9:
                self.reports_menu()
            elif choice == 10:
                self.sample_data_menu()
                self.pause()
            elif choice == 0:
                self.exit_application()

    def exit_application(self):
        """Exit the application with auto-save."""
        console.print(
            "\n[bold yellow]Bạn có muốn lưu dữ liệu trước khi thoát?[/bold yellow]"
        )
        confirm = Confirm.ask("Lưu dữ liệu", default=True)

        if confirm:
            self.persistence.save_data(self)

        console.print(
            Panel(
                "[bold green]Cảm ơn đã sử dụng Hệ thống Xếp hạng Khám bệnh[/bold green]",
                border_style="green",
            )
        )
        self.running = False


def main():
    """Entry point for the CLI application."""
    try:
        cli = HospitalCLI()
        cli.run()
    except KeyboardInterrupt:
        console.print("\n[bold yellow]Đã nhận Ctrl+C, đang thoát...[/bold yellow]")
        sys.exit(0)
    except Exception as e:
        console.print(f"[bold red]Lỗi nghiêm trọng: {e}[/bold red]")
        raise


if __name__ == "__main__":
    main()
