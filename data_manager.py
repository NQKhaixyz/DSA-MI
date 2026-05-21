"""
Hospital Triage Data Manager Module

Provides production-ready data persistence, export/import, auto-save, and backup
functionality for the Hospital Triage System.

Features:
    - JSON serialization with custom encoders/decoders
    - CSV export/import for interoperability
    - Excel export with multiple sheets (requires openpyxl)
    - Auto-save with configurable intervals
    - Timestamped backups before overwrites
    - Data validation and integrity checks

Dependencies:
    - openpyxl (optional, for Excel support)
"""

from __future__ import annotations

import csv
import json
import logging
import os
import shutil
import threading
import time
from collections import deque
from dataclasses import asdict, is_dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import (
    Any,
    Callable,
    Dict,
    List,
    Optional,
    Set,
    Tuple,
    TypeVar,
    Union,
)

from triage_system import Department, Doctor, Patient, TriageSystem

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_BACKUP_DIR = "backups"
DEFAULT_CSV_ENCODING = "utf-8-sig"
DATETIME_FORMAT = "%Y-%m-%dT%H:%M:%S"
JSON_INDENT = 2

T = TypeVar("T")

# ---------------------------------------------------------------------------
# Custom JSON Encoder / Decoder
# ---------------------------------------------------------------------------


class _TriageJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder handling datetime, deque, set, tuple keys, and Patient/Doctor/Department."""

    def default(self, obj: Any) -> Any:
        if isinstance(obj, datetime):
            return {"__type__": "datetime", "value": obj.strftime(DATETIME_FORMAT)}
        if isinstance(obj, timedelta):
            return {
                "__type__": "timedelta",
                "days": obj.days,
                "seconds": obj.seconds,
                "microseconds": obj.microseconds,
            }
        if isinstance(obj, deque):
            return {"__type__": "deque", "value": list(obj)}
        if isinstance(obj, set):
            return {"__type__": "set", "value": list(obj)}
        if isinstance(obj, tuple):
            return {"__type__": "tuple", "value": list(obj)}
        if isinstance(obj, Patient):
            return {"__type__": "Patient", "value": _patient_to_dict(obj)}
        if isinstance(obj, Doctor):
            return {"__type__": "Doctor", "value": _doctor_to_dict(obj)}
        if isinstance(obj, Department):
            return {"__type__": "Department", "value": _department_to_dict(obj)}
        return super().default(obj)

    def iterencode(self, obj: Any, _one_shot: bool = False) -> Any:
        """Override iterencode to handle dicts with tuple keys."""
        obj = self._convert_tuple_keys(obj)
        yield from super().iterencode(obj, _one_shot)

    def _convert_tuple_keys(self, obj: Any) -> Any:
        if isinstance(obj, dict):
            new_dict = {}
            for key, value in obj.items():
                if isinstance(key, tuple):
                    key = "__tuple_key__:" + json.dumps(list(key))
                new_dict[key] = self._convert_tuple_keys(value)
            return new_dict
        if isinstance(obj, list):
            return [self._convert_tuple_keys(item) for item in obj]
        if isinstance(obj, tuple):
            return {"__type__": "tuple", "value": list(obj)}
        if isinstance(obj, set):
            return {"__type__": "set", "value": list(obj)}
        return obj


def _triage_json_object_hook(dct: Dict[str, Any]) -> Any:
    """Object hook that converts tuple_key string markers back to tuple keys and decodes typed objects."""
    # First check if this dict is a typed object itself
    type_tag = dct.get("__type__")
    if type_tag == "datetime":
        return datetime.strptime(dct["value"], DATETIME_FORMAT)
    if type_tag == "timedelta":
        return timedelta(
            days=dct["days"], seconds=dct["seconds"], microseconds=dct["microseconds"]
        )
    if type_tag == "deque":
        return deque(dct["value"])
    if type_tag == "set":
        return set(dct["value"])
    if type_tag == "tuple":
        return tuple(dct["value"])
    if type_tag == "Patient":
        return _dict_to_patient(dct["value"])
    if type_tag == "Doctor":
        return _dict_to_doctor(dct["value"])
    if type_tag == "Department":
        return _dict_to_department(dct["value"])

    # Otherwise, process dict keys and values recursively
    result = {}
    for key, value in dct.items():
        if isinstance(key, str) and key.startswith("__tuple_key__:"):
            key = tuple(json.loads(key[len("__tuple_key__:") :]))
        if isinstance(value, dict):
            value = _triage_json_object_hook(value)
        elif isinstance(value, list):
            value = [
                _triage_json_object_hook(item) if isinstance(item, dict) else item
                for item in value
            ]
        result[key] = value
    return result
    if type_tag == "set":
        return set(dct["value"])
    if type_tag == "Patient":
        return _dict_to_patient(dct["value"])
    if type_tag == "Doctor":
        return _dict_to_doctor(dct["value"])
    if type_tag == "Department":
        return _dict_to_department(dct["value"])
    return dct


# ---------------------------------------------------------------------------
# Object <-> Dict conversion helpers
# ---------------------------------------------------------------------------


def _patient_to_dict(p: Patient) -> Dict[str, Any]:
    return {
        "patient_id": p.patient_id,
        "name": p.name,
        "age": p.age,
        "symptoms": p.symptoms,
        "danger_level": p.danger_level,
        "is_booked": p.is_booked,
        "ticket_number": p.ticket_number,
        "appointment_time": p.appointment_time,
        "doctor_id": p.doctor_id,
        "arrival_time": p.arrival_time,
        "wait_start_time": p.wait_start_time,
        "doctor_ids": p.doctor_ids,
        "priority": p.priority,
    }


def _dict_to_patient(d: Dict[str, Any]) -> Patient:
    p = Patient(
        patient_id=d["patient_id"],
        name=d.get("name", ""),
        age=d.get("age", 0),
        symptoms=d.get("symptoms", ""),
        danger_level=d.get("danger_level", 2),
        is_booked=d.get("is_booked", False),
    )
    p.ticket_number = d.get("ticket_number", "")
    p.appointment_time = d.get("appointment_time")
    p.doctor_id = d.get("doctor_id")
    p.arrival_time = d.get("arrival_time", datetime.now())
    p.wait_start_time = d.get("wait_start_time", datetime.now())
    p.doctor_ids = set(d.get("doctor_ids", []))
    p.priority = d.get("priority", p.danger_level)
    return p


def _doctor_to_dict(d: Doctor) -> Dict[str, Any]:
    return {
        "doctor_id": d.doctor_id,
        "name": d.name,
        "department_id": d.department_id,
        "booked_queue": list(d.booked_queue),
        "current_patient_id": d.current_patient_id,
        "is_idle": d.is_idle,
    }


def _dict_to_doctor(d: Dict[str, Any]) -> Doctor:
    doc = Doctor(
        doctor_id=d["doctor_id"],
        name=d.get("name", ""),
        department_id=d.get("department_id", ""),
    )
    # Rehydrate queue items: they may be Patient objects or string IDs
    raw_queue = d.get("booked_queue", [])
    for item in raw_queue:
        if isinstance(item, dict) and item.get("__type__") == "Patient":
            doc.booked_queue.append(_dict_to_patient(item["value"]))
        else:
            doc.booked_queue.append(item)
    doc.current_patient_id = d.get("current_patient_id")
    doc.is_idle = d.get("is_idle", True)
    return doc


def _department_to_dict(d: Department) -> Dict[str, Any]:
    return {
        "department_id": d.department_id,
        "name": d.name,
        "emergency_queue": list(d.emergency_queue),
        "walk_in_queue": list(d.walk_in_queue),
        "priority_queue": list(d.priority_queue),
        "doctor_ids": list(d.doctor_ids),
    }


def _dict_to_department(d: Dict[str, Any]) -> Department:
    dept = Department(
        department_id=d["department_id"],
        name=d.get("name", ""),
    )
    for key in ("emergency_queue", "walk_in_queue", "priority_queue"):
        for item in d.get(key, []):
            if isinstance(item, dict) and item.get("__type__") == "Patient":
                getattr(dept, key).append(_dict_to_patient(item["value"]))
            else:
                getattr(dept, key).append(item)
    dept.doctor_ids = set(d.get("doctor_ids", []))
    return dept


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------


class DataValidationError(Exception):
    """Raised when loaded data fails integrity checks."""

    pass


def _validate_triage_data(data: Dict[str, Any]) -> None:
    """Validate serialized triage system snapshot."""
    required_keys = {
        "patients",
        "doctors",
        "departments",
        "appointment_slots",
        "scheduled_appointments",
        "walk_in_counter",
        "booked_counter",
        "doctor_patient_graph",
        "patient_doctor_graph",
    }
    missing = required_keys - data.keys()
    if missing:
        raise DataValidationError(f"Missing top-level keys: {missing}")

    # Validate relationships
    doctor_patient_graph: Dict[str, List[str]] = data.get("doctor_patient_graph", {})
    patient_doctor_graph: Dict[str, List[str]] = data.get("patient_doctor_graph", {})

    for doc_id, pids in doctor_patient_graph.items():
        if not isinstance(pids, list):
            raise DataValidationError(
                f"doctor_patient_graph[{doc_id}] must be a list/set of patient IDs"
            )
        for pid in pids:
            if pid not in data.get("patients", {}):
                raise DataValidationError(
                    f"doctor_patient_graph references unknown patient: {pid}"
                )

    for pid, doc_ids in patient_doctor_graph.items():
        if not isinstance(doc_ids, list):
            raise DataValidationError(
                f"patient_doctor_graph[{pid}] must be a list/set of doctor IDs"
            )
        for did in doc_ids:
            if did not in data.get("doctors", {}):
                raise DataValidationError(
                    f"patient_doctor_graph references unknown doctor: {did}"
                )


# ---------------------------------------------------------------------------
# Backup helpers
# ---------------------------------------------------------------------------


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def create_backup(filepath: Union[str, Path], backup_dir: Optional[str] = None) -> str:
    """
    Create a timestamped backup of *filepath* before it is overwritten.

    Args:
        filepath: Original file path.
        backup_dir: Directory to store backups. Defaults to ``backups/`` next to the file.

    Returns:
        Path to the created backup file.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Cannot backup non-existent file: {filepath}")

    backup_dir = Path(backup_dir or filepath.parent / DEFAULT_BACKUP_DIR)
    backup_dir.mkdir(parents=True, exist_ok=True)

    stem = filepath.stem
    suffix = filepath.suffix
    backup_name = f"{stem}_{_timestamp()}{suffix}"
    backup_path = backup_dir / backup_name

    shutil.copy2(filepath, backup_path)
    logger.info("Created backup: %s", backup_path)
    return str(backup_path)


# ---------------------------------------------------------------------------
# JSON Handler
# ---------------------------------------------------------------------------


def save_to_json(triage_system: TriageSystem, filepath: Union[str, Path]) -> str:
    """
    Persist a ``TriageSystem`` snapshot to a JSON file.

    Creates a timestamped backup if the target file already exists.

    Args:
        triage_system: The triage system instance to save.
        filepath: Destination JSON file path.

    Returns:
        The path to the saved file.
    """
    filepath = Path(filepath)
    if filepath.exists():
        create_backup(filepath)

    snapshot = {
        "patients": triage_system.patients,
        "doctors": triage_system.doctors,
        "departments": triage_system.departments,
        "appointment_slots": triage_system.appointment_slots,
        "scheduled_appointments": triage_system.scheduled_appointments,
        "walk_in_counter": triage_system.walk_in_counter,
        "booked_counter": triage_system.booked_counter,
        "doctor_patient_graph": {
            k: list(v) for k, v in triage_system.doctor_patient_graph.items()
        },
        "patient_doctor_graph": {
            k: list(v) for k, v in triage_system.patient_doctor_graph.items()
        },
        "_saved_at": datetime.now().isoformat(),
    }

    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(
                snapshot,
                f,
                cls=_TriageJSONEncoder,
                indent=JSON_INDENT,
                ensure_ascii=False,
            )
        logger.info("Saved triage data to %s", filepath)
    except (OSError, TypeError) as exc:
        logger.error("Failed to save JSON: %s", exc)
        raise

    return str(filepath)


def load_from_json(triage_system: TriageSystem, filepath: Union[str, Path]) -> None:
    """
    Restore a ``TriageSystem`` snapshot from a JSON file.

    Performs validation before mutating the live system.

    Args:
        triage_system: The triage system instance to populate.
        filepath: Source JSON file path.

    Raises:
        FileNotFoundError: If the file does not exist.
        DataValidationError: If the data fails integrity checks.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"JSON file not found: {filepath}")

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw_data = json.load(f)
            data = _triage_json_object_hook(raw_data)
    except json.JSONDecodeError as exc:
        logger.error("Invalid JSON in %s: %s", filepath, exc)
        raise

    _validate_triage_data(data)

    # Clear current state
    triage_system.patients.clear()
    triage_system.doctors.clear()
    triage_system.departments.clear()
    triage_system.appointment_slots.clear()
    triage_system.scheduled_appointments.clear()
    triage_system.doctor_patient_graph.clear()
    triage_system.patient_doctor_graph.clear()

    # Restore primitives
    triage_system.walk_in_counter = data.get("walk_in_counter", 0)
    triage_system.booked_counter = data.get("booked_counter", 0)

    # Restore collections
    for pid, patient in data.get("patients", {}).items():
        triage_system.patients[pid] = (
            patient if isinstance(patient, Patient) else _dict_to_patient(patient)
        )

    for did, doctor in data.get("doctors", {}).items():
        triage_system.doctors[did] = (
            doctor if isinstance(doctor, Doctor) else _dict_to_doctor(doctor)
        )

    for dept_id, dept in data.get("departments", {}).items():
        triage_system.departments[dept_id] = (
            dept if isinstance(dept, Department) else _dict_to_department(dept)
        )

    triage_system.appointment_slots = data.get("appointment_slots", {})
    triage_system.scheduled_appointments = data.get("scheduled_appointments", {})

    triage_system.doctor_patient_graph = {
        k: set(v) for k, v in data.get("doctor_patient_graph", {}).items()
    }
    triage_system.patient_doctor_graph = {
        k: set(v) for k, v in data.get("patient_doctor_graph", {}).items()
    }

    logger.info("Loaded triage data from %s", filepath)


# ---------------------------------------------------------------------------
# CSV Handler
# ---------------------------------------------------------------------------


def _write_csv_rows(
    filepath: Path,
    headers: List[str],
    rows: List[List[Any]],
) -> None:
    with open(filepath, "w", newline="", encoding=DEFAULT_CSV_ENCODING) as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def export_to_csv(
    triage_system: TriageSystem, directory: Union[str, Path]
) -> List[str]:
    """
    Export patients, doctors, and appointments to CSV files inside *directory*.

    Args:
        triage_system: The triage system to export.
        directory: Target directory (created if missing).

    Returns:
        List of exported file paths.
    """
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    exported: List[str] = []

    # Patients
    patients_file = directory / "patients.csv"
    patient_headers = [
        "patient_id",
        "name",
        "age",
        "symptoms",
        "danger_level",
        "is_booked",
        "ticket_number",
        "appointment_time",
        "doctor_id",
        "arrival_time",
        "wait_start_time",
        "priority",
    ]
    patient_rows: List[List[Any]] = []
    for p in triage_system.patients.values():
        patient_rows.append(
            [
                p.patient_id,
                p.name,
                p.age,
                p.symptoms,
                p.danger_level,
                int(p.is_booked),
                p.ticket_number,
                p.appointment_time.strftime(DATETIME_FORMAT)
                if p.appointment_time
                else "",
                p.doctor_id or "",
                p.arrival_time.strftime(DATETIME_FORMAT) if p.arrival_time else "",
                p.wait_start_time.strftime(DATETIME_FORMAT)
                if p.wait_start_time
                else "",
                p.priority,
            ]
        )
    _write_csv_rows(patients_file, patient_headers, patient_rows)
    exported.append(str(patients_file))
    logger.info("Exported %d patients to %s", len(patient_rows), patients_file)

    # Doctors
    doctors_file = directory / "doctors.csv"
    doctor_headers = [
        "doctor_id",
        "name",
        "department_id",
        "current_patient_id",
        "is_idle",
    ]
    doctor_rows: List[List[Any]] = []
    for d in triage_system.doctors.values():
        doctor_rows.append(
            [
                d.doctor_id,
                d.name,
                d.department_id,
                d.current_patient_id or "",
                int(d.is_idle),
            ]
        )
    _write_csv_rows(doctors_file, doctor_headers, doctor_rows)
    exported.append(str(doctors_file))
    logger.info("Exported %d doctors to %s", len(doctor_rows), doctors_file)

    # Appointments
    appointments_file = directory / "appointments.csv"
    appointment_headers = [
        "date",
        "time_slot",
        "doctor_id",
        "patient_id",
        "ticket_number",
        "patient_name",
    ]
    appointment_rows: List[List[Any]] = []
    for slot_key, apps in triage_system.scheduled_appointments.items():
        date, time_slot, doctor_id = slot_key
        for app in apps:
            appointment_rows.append(
                [
                    date,
                    time_slot,
                    doctor_id,
                    app.get("patient_id", ""),
                    app.get("ticket_number", ""),
                    app.get("patient_name", ""),
                ]
            )
    _write_csv_rows(appointments_file, appointment_headers, appointment_rows)
    exported.append(str(appointments_file))
    logger.info(
        "Exported %d appointments to %s", len(appointment_rows), appointments_file
    )

    return exported


def import_from_csv(
    triage_system: TriageSystem, directory: Union[str, Path]
) -> Dict[str, int]:
    """
    Import patients, doctors, and appointments from CSV files in *directory*.

    Existing entries with the same ID are overwritten.

    Args:
        triage_system: The triage system to populate.
        directory: Source directory containing CSV files.

    Returns:
        Dictionary with counts of imported items per type.
    """
    directory = Path(directory)
    counts: Dict[str, int] = {"patients": 0, "doctors": 0, "appointments": 0}

    def _parse_datetime(value: str) -> Optional[datetime]:
        return datetime.strptime(value, DATETIME_FORMAT) if value else None

    # Patients
    patients_file = directory / "patients.csv"
    if patients_file.exists():
        with open(patients_file, "r", encoding=DEFAULT_CSV_ENCODING, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                p = Patient(
                    patient_id=row["patient_id"],
                    name=row.get("name", ""),
                    age=int(row.get("age", 0) or 0),
                    symptoms=row.get("symptoms", ""),
                    danger_level=int(row.get("danger_level", 2) or 2),
                    is_booked=bool(int(row.get("is_booked", 0) or 0)),
                )
                p.ticket_number = row.get("ticket_number", "")
                p.appointment_time = _parse_datetime(row.get("appointment_time", ""))
                p.doctor_id = row.get("doctor_id") or None
                p.arrival_time = (
                    _parse_datetime(row.get("arrival_time", "")) or datetime.now()
                )
                p.wait_start_time = (
                    _parse_datetime(row.get("wait_start_time", "")) or datetime.now()
                )
                p.priority = int(row.get("priority", p.danger_level) or p.danger_level)
                triage_system.patients[p.patient_id] = p
                counts["patients"] += 1
        logger.info("Imported %d patients from %s", counts["patients"], patients_file)

    # Doctors
    doctors_file = directory / "doctors.csv"
    if doctors_file.exists():
        with open(doctors_file, "r", encoding=DEFAULT_CSV_ENCODING, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                doc = Doctor(
                    doctor_id=row["doctor_id"],
                    name=row.get("name", ""),
                    department_id=row.get("department_id", ""),
                )
                doc.current_patient_id = row.get("current_patient_id") or None
                doc.current_patient_id = (
                    doc.current_patient_id if doc.current_patient_id != "" else None
                )
                doc.is_idle = bool(int(row.get("is_idle", 1) or 1))
                triage_system.doctors[doc.doctor_id] = doc
                counts["doctors"] += 1
        logger.info("Imported %d doctors from %s", counts["doctors"], doctors_file)

    # Appointments
    appointments_file = directory / "appointments.csv"
    if appointments_file.exists():
        with open(
            appointments_file, "r", encoding=DEFAULT_CSV_ENCODING, newline=""
        ) as f:
            reader = csv.DictReader(f)
            for row in reader:
                slot_key = (row["date"], row["time_slot"], row["doctor_id"])
                if slot_key not in triage_system.scheduled_appointments:
                    triage_system.scheduled_appointments[slot_key] = []
                triage_system.scheduled_appointments[slot_key].append(
                    {
                        "patient_id": row.get("patient_id", ""),
                        "ticket_number": row.get("ticket_number", ""),
                        "date": row["date"],
                        "time_slot": row["time_slot"],
                        "doctor_id": row["doctor_id"],
                        "patient_name": row.get("patient_name", ""),
                    }
                )
                # Update slot counter
                triage_system.appointment_slots[slot_key] = len(
                    triage_system.scheduled_appointments[slot_key]
                )
                counts["appointments"] += 1
        logger.info(
            "Imported %d appointments from %s",
            counts["appointments"],
            appointments_file,
        )

    return counts


# ---------------------------------------------------------------------------
# Excel Handler
# ---------------------------------------------------------------------------


def export_to_excel(triage_system: TriageSystem, filepath: Union[str, Path]) -> str:
    """
    Export triage data to an Excel workbook with multiple sheets.

    Requires ``openpyxl`` to be installed.

    Sheets created:
        - Patients
        - Doctors
        - Departments
        - Appointments
        - Appointment Slots

    Args:
        triage_system: The triage system to export.
        filepath: Destination ``.xlsx`` file path.

    Returns:
        The path to the exported file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from exc

    filepath = Path(filepath)
    if filepath.exists():
        create_backup(filepath)

    wb = openpyxl.Workbook()

    # Helper to create a styled sheet
    def _make_sheet(name: str, headers: List[str], rows: List[List[Any]]) -> None:
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)

    # Patients sheet
    patient_headers = [
        "patient_id",
        "name",
        "age",
        "symptoms",
        "danger_level",
        "is_booked",
        "ticket_number",
        "appointment_time",
        "doctor_id",
        "arrival_time",
        "wait_start_time",
        "priority",
    ]
    patient_rows: List[List[Any]] = []
    for p in triage_system.patients.values():
        patient_rows.append(
            [
                p.patient_id,
                p.name,
                p.age,
                p.symptoms,
                p.danger_level,
                p.is_booked,
                p.ticket_number,
                p.appointment_time.strftime(DATETIME_FORMAT)
                if p.appointment_time
                else "",
                p.doctor_id or "",
                p.arrival_time.strftime(DATETIME_FORMAT) if p.arrival_time else "",
                p.wait_start_time.strftime(DATETIME_FORMAT)
                if p.wait_start_time
                else "",
                p.priority,
            ]
        )
    _make_sheet("Patients", patient_headers, patient_rows)

    # Doctors sheet
    doctor_headers = [
        "doctor_id",
        "name",
        "department_id",
        "current_patient_id",
        "is_idle",
    ]
    doctor_rows: List[List[Any]] = []
    for d in triage_system.doctors.values():
        doctor_rows.append(
            [
                d.doctor_id,
                d.name,
                d.department_id,
                d.current_patient_id or "",
                d.is_idle,
            ]
        )
    _make_sheet("Doctors", doctor_headers, doctor_rows)

    # Departments sheet
    dept_headers = [
        "department_id",
        "name",
        "doctor_ids",
        "emergency_queue_len",
        "walk_in_queue_len",
        "priority_queue_len",
    ]
    dept_rows: List[List[Any]] = []
    for d in triage_system.departments.values():
        dept_rows.append(
            [
                d.department_id,
                d.name,
                ", ".join(d.doctor_ids),
                len(d.emergency_queue),
                len(d.walk_in_queue),
                len(d.priority_queue),
            ]
        )
    _make_sheet("Departments", dept_headers, dept_rows)

    # Appointments sheet
    app_headers = [
        "date",
        "time_slot",
        "doctor_id",
        "patient_id",
        "ticket_number",
        "patient_name",
    ]
    app_rows: List[List[Any]] = []
    for slot_key, apps in triage_system.scheduled_appointments.items():
        date, time_slot, doctor_id = slot_key
        for app in apps:
            app_rows.append(
                [
                    date,
                    time_slot,
                    doctor_id,
                    app.get("patient_id", ""),
                    app.get("ticket_number", ""),
                    app.get("patient_name", ""),
                ]
            )
    _make_sheet("Appointments", app_headers, app_rows)

    # Appointment Slots sheet
    slot_headers = ["date", "time_slot", "doctor_id", "booked_count", "max_capacity"]
    slot_rows: List[List[Any]] = []
    for slot_key, count in triage_system.appointment_slots.items():
        date, time_slot, doctor_id = slot_key
        slot_rows.append([date, time_slot, doctor_id, count, 4])
    _make_sheet("Appointment Slots", slot_headers, slot_rows)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(filepath)
    logger.info("Exported triage data to Excel: %s", filepath)
    return str(filepath)


# ---------------------------------------------------------------------------
# Auto-save
# ---------------------------------------------------------------------------


class AutoSaveManager:
    """
    Background auto-save manager for ``TriageSystem``.

    Supports interval-based saving and optional immediate save on every change
    (via a lightweight dirty flag checked by the background thread).
    """

    def __init__(
        self,
        triage_system: TriageSystem,
        filepath: Union[str, Path],
        interval_minutes: float = 5.0,
        save_on_change: bool = False,
    ):
        self.triage_system = triage_system
        self.filepath = Path(filepath)
        self.interval_seconds = interval_minutes * 60.0
        self.save_on_change = save_on_change
        self._dirty = False
        self._stop_event = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._lock = threading.Lock()

    def _save_loop(self) -> None:
        while not self._stop_event.is_set():
            save_now = False
            with self._lock:
                if self._dirty:
                    save_now = True
                    self._dirty = False
            if save_now:
                try:
                    save_to_json(self.triage_system, self.filepath)
                except Exception as exc:
                    logger.error("Auto-save failed: %s", exc)
            self._stop_event.wait(timeout=self.interval_seconds)

    def start(self) -> None:
        """Start the background auto-save thread."""
        if self._thread is not None and self._thread.is_alive():
            logger.warning("Auto-save thread already running")
            return
        self._stop_event.clear()
        self._thread = threading.Thread(
            target=self._save_loop, daemon=True, name="TriageAutoSave"
        )
        self._thread.start()
        logger.info(
            "Auto-save started: %s every %.1f minutes",
            self.filepath,
            self.interval_seconds / 60,
        )

    def stop(self) -> None:
        """Stop the background auto-save thread."""
        self._stop_event.set()
        if self._thread is not None:
            self._thread.join(timeout=2.0)
            self._thread = None
        logger.info("Auto-save stopped")

    def mark_dirty(self) -> None:
        """Mark the system as changed so the next interval will trigger a save."""
        if self.save_on_change:
            with self._lock:
                self._dirty = True

    def force_save(self) -> str:
        """Immediately save regardless of dirty state."""
        path = save_to_json(self.triage_system, self.filepath)
        with self._lock:
            self._dirty = False
        return path


def auto_save(
    triage_system: TriageSystem,
    filepath: Union[str, Path],
    interval_minutes: float = 5.0,
    save_on_change: bool = False,
) -> AutoSaveManager:
    """
    Convenience factory that creates, starts, and returns an ``AutoSaveManager``.

    Args:
        triage_system: The triage system to auto-save.
        filepath: Target JSON file path.
        interval_minutes: Minutes between auto-save attempts.
        save_on_change: If ``True``, any call to ``mark_dirty()`` will cause a save
            on the next interval tick.

    Returns:
        The running ``AutoSaveManager`` instance.  Call ``.stop()`` when done.
    """
    manager = AutoSaveManager(
        triage_system=triage_system,
        filepath=filepath,
        interval_minutes=interval_minutes,
        save_on_change=save_on_change,
    )
    manager.start()
    return manager


# ---------------------------------------------------------------------------
# Module-level convenience
# ---------------------------------------------------------------------------

__all__ = [
    "save_to_json",
    "load_from_json",
    "export_to_csv",
    "import_from_csv",
    "export_to_excel",
    "auto_save",
    "AutoSaveManager",
    "create_backup",
    "DataValidationError",
]
