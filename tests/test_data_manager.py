import pytest
import json
import csv
import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from triage_system import TriageSystem, Patient, Doctor, Department
from data_manager import (
    save_to_json,
    load_from_json,
    export_to_csv,
    import_from_csv,
    export_to_excel,
    create_backup,
    DataValidationError,
    _TriageJSONEncoder,
    _triage_json_object_hook,
    _patient_to_dict,
    _dict_to_patient,
    _doctor_to_dict,
    _dict_to_doctor,
    _department_to_dict,
    _dict_to_department,
    AutoSaveManager,
)


@pytest.fixture
def temp_dir(tmp_path):
    return tmp_path


@pytest.fixture
def populated_system():
    ts = TriageSystem()
    ts.departments["DEPT01"] = Department("DEPT01", "Emergency")
    ts.departments["DEPT02"] = Department("DEPT02", "Cardiology")
    ts.doctors["D001"] = Doctor("D001", "Dr. Smith", "DEPT01")
    ts.doctors["D002"] = Doctor("D002", "Dr. Jones", "DEPT02")
    ts.departments["DEPT01"].doctor_ids.add("D001")
    ts.departments["DEPT02"].doctor_ids.add("D002")

    ts.patients["P001"] = Patient("P001", "John", age=30, danger_level=3)
    ts.patients["P002"] = Patient(
        "P002", "Jane", age=25, danger_level=2, is_booked=True
    )
    ts.patients["P002"].doctor_id = "D001"
    ts.patients["P002"].appointment_time = datetime(2024, 1, 15, 9, 0)

    ts.book_appointment(
        {"patient_id": "P003", "name": "Bob"}, "2024-01-15", "09:00-10:00", "D001"
    )
    return ts


class TestJSONSaveLoad:
    def test_save_to_json_creates_file(self, populated_system, temp_dir):
        filepath = temp_dir / "triage.json"
        result = save_to_json(populated_system, filepath)
        assert Path(result).exists()

    def test_save_to_json_content(self, populated_system, temp_dir):
        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert "patients" in data
        assert "doctors" in data
        assert "departments" in data

    def test_load_from_json(self, populated_system, temp_dir):
        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)

        new_system = TriageSystem()
        load_from_json(new_system, filepath)
        assert "P001" in new_system.patients
        assert "D001" in new_system.doctors
        assert "DEPT01" in new_system.departments

    def test_load_preserves_patient_data(self, populated_system, temp_dir):
        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)

        new_system = TriageSystem()
        load_from_json(new_system, filepath)
        patient = new_system.patients["P001"]
        assert patient.name == "John"
        assert patient.age == 30
        assert patient.danger_level == 3

    def test_round_trip_doctor_queue(self, populated_system, temp_dir):
        patient = Patient("P004", "Booked", danger_level=2, is_booked=True)
        populated_system.add_to_booked_queue(
            "D001", patient, datetime(2024, 1, 15, 10, 0)
        )

        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)

        new_system = TriageSystem()
        load_from_json(new_system, filepath)
        assert len(new_system.doctors["D001"].booked_queue) == 1

    def test_round_trip_department_queue(self, populated_system, temp_dir):
        patient = Patient("P005", "Emergency", danger_level=1)
        populated_system.add_to_emergency_queue("DEPT01", patient)

        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)

        new_system = TriageSystem()
        load_from_json(new_system, filepath)
        assert len(new_system.departments["DEPT01"].emergency_queue) == 1

    def test_load_invalid_file(self, temp_dir):
        new_system = TriageSystem()
        with pytest.raises(FileNotFoundError):
            load_from_json(new_system, temp_dir / "nonexistent.json")

    def test_load_invalid_json(self, temp_dir):
        filepath = temp_dir / "invalid.json"
        with open(filepath, "w") as f:
            f.write("invalid json")
        new_system = TriageSystem()
        with pytest.raises(Exception):
            load_from_json(new_system, filepath)


class TestCSVExportImport:
    def test_export_to_csv_creates_files(self, populated_system, temp_dir):
        exported = export_to_csv(populated_system, temp_dir)
        assert len(exported) == 3
        assert any("patients.csv" in f for f in exported)
        assert any("doctors.csv" in f for f in exported)
        assert any("appointments.csv" in f for f in exported)

    def test_export_patient_content(self, populated_system, temp_dir):
        export_to_csv(populated_system, temp_dir)
        with open(temp_dir / "patients.csv", "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 3
        assert any(row["patient_id"] == "P001" for row in rows)

    def test_import_from_csv(self, populated_system, temp_dir):
        export_to_csv(populated_system, temp_dir)

        new_system = TriageSystem()
        counts = import_from_csv(new_system, temp_dir)
        assert counts["patients"] == 3
        assert counts["doctors"] == 2
        assert counts["appointments"] == 1

    def test_import_preserves_patient_data(self, populated_system, temp_dir):
        export_to_csv(populated_system, temp_dir)

        new_system = TriageSystem()
        import_from_csv(new_system, temp_dir)
        patient = new_system.patients["P001"]
        assert patient.name == "John"
        assert patient.age == 30

    def test_import_doctor_state(self, populated_system, temp_dir):
        populated_system.doctors["D001"].current_patient_id = "P001"
        populated_system.doctors["D001"].is_idle = False

        export_to_csv(populated_system, temp_dir)

        new_system = TriageSystem()
        import_from_csv(new_system, temp_dir)
        assert new_system.doctors["D001"].current_patient_id == "P001"
        assert new_system.doctors["D001"].is_idle is False

    def test_import_missing_files(self, temp_dir):
        new_system = TriageSystem()
        counts = import_from_csv(new_system, temp_dir)
        assert counts == {"patients": 0, "doctors": 0, "appointments": 0}

    def test_csv_round_trip_appointment_slots(self, populated_system, temp_dir):
        export_to_csv(populated_system, temp_dir)

        new_system = TriageSystem()
        import_from_csv(new_system, temp_dir)
        slot_key = ("2024-01-15", "09:00-10:00", "D001")
        assert slot_key in new_system.scheduled_appointments
        assert new_system.appointment_slots.get(slot_key, 0) == 1


class TestExcelExport:
    def test_export_to_excel_creates_file(self, populated_system, temp_dir):
        pytest.importorskip("openpyxl")
        filepath = temp_dir / "triage.xlsx"
        result = export_to_excel(populated_system, filepath)
        assert Path(result).exists()

    def test_export_to_excel_sheets(self, populated_system, temp_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        filepath = temp_dir / "triage.xlsx"
        export_to_excel(populated_system, filepath)

        wb = load_workbook(filepath)
        expected_sheets = {
            "Patients",
            "Doctors",
            "Departments",
            "Appointments",
            "Appointment Slots",
        }
        assert set(wb.sheetnames) == expected_sheets

    def test_export_to_excel_patient_data(self, populated_system, temp_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        filepath = temp_dir / "triage.xlsx"
        export_to_excel(populated_system, filepath)

        wb = load_workbook(filepath)
        ws = wb["Patients"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "patient_id"
        assert any(row[0] == "P001" for row in rows[1:])

    def test_export_to_excel_without_openpyxl(
        self, populated_system, temp_dir, monkeypatch
    ):
        import sys

        monkeypatch.setitem(sys.modules, "openpyxl", None)

        filepath = temp_dir / "triage.xlsx"
        with pytest.raises(ImportError):
            export_to_excel(populated_system, filepath)


class TestBackupCreation:
    def test_create_backup(self, temp_dir):
        filepath = temp_dir / "test.json"
        with open(filepath, "w") as f:
            f.write('{"test": true}')

        backup_path = create_backup(filepath)
        assert Path(backup_path).exists()
        assert "test_" in backup_path
        assert backup_path.endswith(".json")

    def test_create_backup_custom_dir(self, temp_dir):
        filepath = temp_dir / "test.json"
        with open(filepath, "w") as f:
            f.write('{"test": true}')

        backup_dir = temp_dir / "custom_backups"
        backup_path = create_backup(filepath, str(backup_dir))
        assert Path(backup_path).exists()
        assert backup_dir.exists()

    def test_create_backup_nonexistent_file(self, temp_dir):
        with pytest.raises(FileNotFoundError):
            create_backup(temp_dir / "nonexistent.json")

    def test_save_to_json_creates_backup(self, populated_system, temp_dir):
        filepath = temp_dir / "triage.json"
        save_to_json(populated_system, filepath)

        # Modify and save again
        populated_system.patients["P006"] = Patient("P006", "New")
        save_to_json(populated_system, filepath)

        backup_dir = temp_dir / "backups"
        assert backup_dir.exists()
        backups = list(backup_dir.glob("triage_*.json"))
        assert len(backups) >= 1


class TestJSONEncoderDecoder:
    def test_datetime_encoding(self):
        dt = datetime(2024, 1, 15, 10, 0, 0)
        encoded = json.dumps({"dt": dt}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        assert decoded["dt"] == dt

    def test_timedelta_encoding(self):
        td = timedelta(hours=2, minutes=30)
        encoded = json.dumps({"td": td}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        assert decoded["td"] == td

    def test_set_encoding(self):
        s = {1, 2, 3}
        encoded = json.dumps({"s": s}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        assert decoded["s"] == s

    def test_deque_encoding(self):
        d = ["a", "b", "c"]
        from collections import deque

        dq = deque(d)
        encoded = json.dumps({"dq": dq}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        assert list(decoded["dq"]) == list(dq)

    def test_patient_round_trip(self):
        patient = Patient("P001", "John", age=30, danger_level=3)
        patient.ticket_number = "W-001"
        patient.doctor_ids = {"D001"}

        encoded = json.dumps({"patient": patient}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        restored = decoded["patient"]
        assert restored.patient_id == "P001"
        assert restored.name == "John"
        assert restored.age == 30

    def test_doctor_round_trip(self):
        doctor = Doctor("D001", "Dr. Smith", "DEPT01")
        doctor.current_patient_id = "P001"
        doctor.is_idle = False

        encoded = json.dumps({"doctor": doctor}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        restored = decoded["doctor"]
        assert restored.doctor_id == "D001"
        assert restored.current_patient_id == "P001"
        assert restored.is_idle is False

    def test_department_round_trip(self):
        dept = Department("DEPT01", "Emergency")
        dept.doctor_ids = {"D001", "D002"}

        encoded = json.dumps({"dept": dept}, cls=_TriageJSONEncoder)
        decoded = json.loads(encoded, object_hook=_triage_json_object_hook)
        restored = decoded["dept"]
        assert restored.department_id == "DEPT01"
        assert restored.doctor_ids == {"D001", "D002"}


class TestValidation:
    def test_validate_triage_data_missing_keys(self):
        with pytest.raises(DataValidationError):
            from data_manager import _validate_triage_data

            _validate_triage_data({"patients": {}, "doctors": {}})

    def test_validate_invalid_doctor_patient_graph(self):
        with pytest.raises(DataValidationError):
            from data_manager import _validate_triage_data

            data = {
                "patients": {},
                "doctors": {"D001": {}},
                "departments": {},
                "appointment_slots": {},
                "scheduled_appointments": {},
                "walk_in_counter": 0,
                "booked_counter": 0,
                "doctor_patient_graph": {"D001": ["P001"]},
                "patient_doctor_graph": {},
            }
            _validate_triage_data(data)


class TestAutoSaveManager:
    def test_autosave_manager_creation(self, populated_system, temp_dir):
        filepath = temp_dir / "autosave.json"
        manager = AutoSaveManager(populated_system, filepath, interval_minutes=0.1)
        assert manager.filepath == filepath
        assert manager.interval_seconds == 6.0

    def test_autosave_start_stop(self, populated_system, temp_dir):
        filepath = temp_dir / "autosave.json"
        manager = AutoSaveManager(populated_system, filepath, interval_minutes=0.1)
        manager.start()
        assert manager._thread is not None
        assert manager._thread.is_alive()
        manager.stop()
        assert manager._thread is None

    def test_autosave_force_save(self, populated_system, temp_dir):
        filepath = temp_dir / "autosave.json"
        manager = AutoSaveManager(populated_system, filepath)
        result = manager.force_save()
        assert Path(result).exists()

    def test_autosave_mark_dirty(self, populated_system, temp_dir):
        filepath = temp_dir / "autosave.json"
        manager = AutoSaveManager(populated_system, filepath, save_on_change=True)
        manager.mark_dirty()
        assert manager._dirty is True
